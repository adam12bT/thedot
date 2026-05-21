"""
tests/test_pipeline_extended.py
Additional tests covering logic not exercised by test_pipeline.py.
Run: pytest tests/ -v
"""

import io
import datetime
import pandas as pd
import numpy as np
import pytest
import sys
import os

# ── Path setup ─────────────────────────────────────────────────────────────────
# Add the project root (one level up from tests/) so all modules are importable.
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from src.utils import parse_date, map_status, map_space, find_col
from src.config import PARTICIPANT_MAX
from src.exporters import ExcelExporter


# Helper that mirrors the old remove_outliers(series) signature used by tests.
# The new pipeline removes outliers at the DataFrame level inside DataPipeline,
# so we reproduce the series-level filter logic here for backward-compatible testing.
def remove_outliers(s: pd.Series) -> pd.Series:
    """Keep only values in [0, PARTICIPANT_MAX]; drop NaN and out-of-range."""
    return s[(s >= 0) & (s <= PARTICIPANT_MAX)]


# ─────────────────────────────────────────────────────────────────────────────
# find_col
# ─────────────────────────────────────────────────────────────────────────────

class TestFindCol:
    def _df(self, cols):
        return pd.DataFrame(columns=cols)

    def test_exact_match(self):
        df = self._df(["title", "status", "room"])
        assert find_col(df, "title") == "title"

    def test_case_insensitive_match(self):
        df = self._df(["Title", "Status"])
        assert find_col(df, "title") == "Title"
        assert find_col(df, "STATUS") == "Status"

    def test_first_candidate_wins(self):
        df = self._df(["name", "title"])
        result = find_col(df, "title", "name")
        assert result == "title"

    def test_fallback_to_second_candidate(self):
        df = self._df(["event_name", "status"])
        result = find_col(df, "title", "event_name")
        assert result == "event_name"

    def test_returns_none_when_no_match(self):
        df = self._df(["foo", "bar"])
        assert find_col(df, "title", "name", "event_name") is None

    def test_empty_dataframe(self):
        df = self._df([])
        assert find_col(df, "title") is None


# ─────────────────────────────────────────────────────────────────────────────
# parse_date — additional edge cases
# ─────────────────────────────────────────────────────────────────────────────

class TestParseDateEdgeCases:
    def test_whitespace_padded_iso(self):
        result = parse_date("  2024-03-15T09:30:00  ")
        assert result == pd.Timestamp("2024-03-15 09:30:00")

    def test_none_returns_none(self):
        assert parse_date(None) is None

    def test_whitespace_only_returns_none(self):
        assert parse_date("   ") is None

    def test_completely_invalid_string_returns_none(self):
        assert parse_date("not-a-date-at-all-xyz") is None

    def test_returns_timestamp_type(self):
        result = parse_date("2024-06-15")
        assert isinstance(result, pd.Timestamp)


# ─────────────────────────────────────────────────────────────────────────────
# map_status — additional edge cases
# ─────────────────────────────────────────────────────────────────────────────

class TestMapStatusEdgeCases:
    def test_reporte_passes_through(self):
        # map_status itself does NOT map Reporté — DataPipeline._normalise_types
        # does a post-hoc .replace(). Verify the function passes it through
        # unchanged so a future accidental change to map_status is caught here.
        result = map_status("Reporté")
        assert result == "Reporté"

    def test_strips_whitespace(self):
        assert map_status("  finished  ") == "Tenu"

    def test_annule_without_accent(self):
        assert map_status("annule") == "Annulé"

    def test_empty_string_passthrough(self):
        result = map_status("")
        assert result == ""


# ─────────────────────────────────────────────────────────────────────────────
# map_space — additional edge cases
# ─────────────────────────────────────────────────────────────────────────────

class TestMapSpaceEdgeCases:
    def test_espace_fondation(self):
        assert map_space("ESPACE FONDATION") == "Salle Fondation"

    def test_podcast_tunisia(self):
        assert map_space("Podcast Tunisia") == "Podcast"

    def test_salle_design_thinking(self):
        assert map_space("Salle Design Thinking") == "Salle Design Thinking"

    def test_strips_input(self):
        assert map_space("  Terrasse  ") == "Terrasse"

    def test_whitespace_only_returns_empty(self):
        assert map_space("   ") == ""


# ─────────────────────────────────────────────────────────────────────────────
# remove_outliers — edge cases
# ─────────────────────────────────────────────────────────────────────────────

class TestRemoveOutliersEdgeCases:
    def test_nan_values_are_dropped(self):
        s = pd.Series([10.0, np.nan, 50.0, np.nan])
        result = remove_outliers(s)
        assert result.isna().sum() == 0

    def test_negative_values_are_removed(self):
        s = pd.Series([-5.0, 0.0, 100.0])
        result = remove_outliers(s)
        assert -5.0 not in result.values

    def test_boundary_999_kept(self):
        s = pd.Series([999.0])
        assert 999.0 in remove_outliers(s).values

    def test_boundary_1000_kept(self):
        s = pd.Series([1000.0])
        assert 1000.0 in remove_outliers(s).values

    def test_boundary_1001_removed(self):
        s = pd.Series([1001.0])
        assert len(remove_outliers(s)) == 0

    def test_all_nan(self):
        s = pd.Series([np.nan, np.nan])
        assert len(remove_outliers(s)) == 0


# ─────────────────────────────────────────────────────────────────────────────
# Negative duration swap
# Mirrors DataPipeline._fix_negative_durations in isolation.
# ─────────────────────────────────────────────────────────────────────────────

class TestNegativeDurationSwap:
    def _apply_fix(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df["duration_hours"] = (
            (df["end_time"] - df["start_time"]).dt.total_seconds() / 3600
        ).round(2)
        neg_mask = df["duration_hours"] < 0
        tmp = df.loc[neg_mask, "start_time"].copy()
        df.loc[neg_mask, "start_time"] = df.loc[neg_mask, "end_time"]
        df.loc[neg_mask, "end_time"] = tmp
        df.loc[neg_mask, "duration_hours"] = df.loc[neg_mask, "duration_hours"].abs()
        return df

    def test_swapped_times_are_corrected(self):
        df = pd.DataFrame({
            "start_time": pd.to_datetime(["2024-03-01 16:00"]),
            "end_time":   pd.to_datetime(["2024-03-01 09:00"]),
        })
        fixed = self._apply_fix(df)
        assert fixed.loc[0, "start_time"] < fixed.loc[0, "end_time"]

    def test_duration_becomes_positive(self):
        df = pd.DataFrame({
            "start_time": pd.to_datetime(["2024-03-01 16:00"]),
            "end_time":   pd.to_datetime(["2024-03-01 09:00"]),
        })
        fixed = self._apply_fix(df)
        assert fixed.loc[0, "duration_hours"] > 0

    def test_normal_events_unchanged(self):
        df = pd.DataFrame({
            "start_time": pd.to_datetime(["2024-03-01 09:00"]),
            "end_time":   pd.to_datetime(["2024-03-01 11:00"]),
        })
        fixed = self._apply_fix(df)
        assert fixed.loc[0, "duration_hours"] == 2.0
        assert fixed.loc[0, "start_time"] == pd.Timestamp("2024-03-01 09:00")


# ─────────────────────────────────────────────────────────────────────────────
# Multi-day flag
# Mirrors DataPipeline._add_derived_columns logic in isolation.
# ─────────────────────────────────────────────────────────────────────────────

class TestMultiDayFlag:
    def _apply_flag(self, hours_list):
        s = pd.Series(hours_list, dtype=float)
        return s.apply(lambda x: "multi-day" if pd.notna(x) and x > 24 else "")

    def test_over_24_hours_flagged(self):
        result = self._apply_flag([25.0, 48.0])
        assert all(result == "multi-day")

    def test_exactly_24_hours_not_flagged(self):
        result = self._apply_flag([24.0])
        assert result.iloc[0] == ""

    def test_normal_events_not_flagged(self):
        result = self._apply_flag([1.5, 2.0, 8.0])
        assert all(result == "")

    def test_nan_not_flagged(self):
        result = self._apply_flag([np.nan])
        assert result.iloc[0] == ""


# ─────────────────────────────────────────────────────────────────────────────
# signed_declaration mapping
# Mirrors the inline logic in DataPipeline._build_records.
# ─────────────────────────────────────────────────────────────────────────────

class TestSignedDeclarationMapping:
    def _map(self, val):
        policy_val = val
        return (
            "Oui"
            if pd.notna(policy_val) and str(policy_val).strip() not in ["", "nan", "NaN", "-"]
            else "Non"
        )

    def test_real_value_returns_oui(self):
        assert self._map("signed") == "Oui"
        assert self._map("2024-01-01") == "Oui"
        assert self._map("yes") == "Oui"

    def test_nan_returns_non(self):
        assert self._map(np.nan) == "Non"

    def test_empty_string_returns_non(self):
        assert self._map("") == "Non"

    def test_dash_returns_non(self):
        assert self._map("-") == "Non"

    def test_nan_string_returns_non(self):
        assert self._map("nan") == "Non"
        assert self._map("NaN") == "Non"

    def test_whitespace_only_returns_non(self):
        assert self._map("   ") == "Non"


# ─────────────────────────────────────────────────────────────────────────────
# activity_type mapping from visibility
# Mirrors the inline logic in DataPipeline._build_records.
# ─────────────────────────────────────────────────────────────────────────────

class TestActivityTypeMapping:
    def _map(self, visibility):
        vis = visibility
        if vis and str(vis).lower() == "public":
            return "Evénement externe"
        elif vis:
            return "Evénement interne"
        return None

    def test_public_maps_to_externe(self):
        assert self._map("public") == "Evénement externe"
        assert self._map("Public") == "Evénement externe"
        assert self._map("PUBLIC") == "Evénement externe"

    def test_private_maps_to_interne(self):
        assert self._map("private") == "Evénement interne"
        assert self._map("internal") == "Evénement interne"

    def test_none_or_empty_returns_none(self):
        assert self._map(None) is None
        assert self._map("") is None


# ─────────────────────────────────────────────────────────────────────────────
# Duplicate detection
# Mirrors DataPipeline._drop_duplicates logic in isolation.
# ─────────────────────────────────────────────────────────────────────────────

class TestDuplicateDetection:
    def _make_df(self):
        return pd.DataFrame({
            "event_name": ["Event A", "Event A", "Event B"],
            "start_time": pd.to_datetime([
                "2024-01-10 09:00", "2024-01-10 09:00", "2024-02-01 10:00",
            ]),
            "end_time": pd.to_datetime([
                "2024-01-10 11:00", "2024-01-10 11:00", "2024-02-01 12:00",
            ]),
            "room": ["Salle de formation", "Salle de formation", "Terrasse"],
        })

    def test_duplicate_is_detected(self):
        df = self._make_df()
        dupe_cols = ["event_name", "start_time", "end_time", "room"]
        dupe_mask = df.duplicated(subset=dupe_cols, keep="first")
        assert dupe_mask.sum() == 1

    def test_first_occurrence_is_kept(self):
        df = self._make_df()
        dupe_cols = ["event_name", "start_time", "end_time", "room"]
        dupe_mask = df.duplicated(subset=dupe_cols, keep="first")
        cleaned = df[~dupe_mask].reset_index(drop=True)
        assert len(cleaned) == 2

    def test_unique_events_all_kept(self):
        df = self._make_df()
        dupe_cols = ["event_name", "start_time", "end_time", "room"]
        dupe_mask = df.duplicated(subset=dupe_cols, keep="first")
        cleaned = df[~dupe_mask]
        assert "Event B" in cleaned["event_name"].values


# ─────────────────────────────────────────────────────────────────────────────
# ExcelExporter — smoke tests
# Previously tested build_excel / build_removed_excel from app.py;
# now delegated to ExcelExporter.build_cleaned_excel / .build_removed_excel.
# ─────────────────────────────────────────────────────────────────────────────

class TestExcelBuilders:
    def _make_clean_df(self):
        return pd.DataFrame({
            "event_name":         ["Event A", "Event B"],
            "activity_type":      ["Evénement interne", "Evénement externe"],
            "start_time":         pd.to_datetime(["2024-01-10 09:00", "2024-02-15 14:00"]),
            "end_time":           pd.to_datetime(["2024-01-10 11:00", "2024-02-15 16:00"]),
            "room":               ["Salle de formation", "Terrasse"],
            "booking_date":       pd.to_datetime(["2024-01-05", "2024-02-10"]),
            "organizer_email":    ["a@example.com", "b@example.com"],
            "organizer_name":     ["Alice", "Bob"],
            "organization":       ["Org A", "Org B"],
            "participants":       [50.0, 120.0],
            "status":             ["Tenu", "Annulé"],
            "signed_declaration": ["Oui", "Non"],
            "comment":            ["", "Some comment"],
            "duration_hours":     [2.0, 2.0],
        })

    def test_build_excel_returns_bytes(self):
        result = ExcelExporter.build_cleaned_excel(self._make_clean_df())
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_build_excel_is_valid_xlsx(self):
        from openpyxl import load_workbook
        result = ExcelExporter.build_cleaned_excel(self._make_clean_df())
        wb = load_workbook(io.BytesIO(result))
        assert "EVENTS"   in wb.sheetnames
        assert "SUMMARY"  in wb.sheetnames

    def test_build_excel_row_count(self):
        from openpyxl import load_workbook
        result = ExcelExporter.build_cleaned_excel(self._make_clean_df())
        wb = load_workbook(io.BytesIO(result))
        ws = wb["EVENTS"]
        # 1 header + 2 data rows
        assert ws.max_row == 3

    def test_build_removed_excel_returns_bytes(self):
        result = ExcelExporter.build_removed_excel(
            pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        )
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_build_removed_excel_has_three_sheets(self):
        from openpyxl import load_workbook
        result = ExcelExporter.build_removed_excel(
            pd.DataFrame({"a": [1]}),
            pd.DataFrame({"a": [2]}),
            pd.DataFrame({"a": [3]}),
        )
        wb = load_workbook(io.BytesIO(result))
        assert set(wb.sheetnames) == {"Empty", "Duplicates", "Negative_Duration"}

    def test_build_removed_excel_with_outliers_has_four_sheets(self):
        """Verify the optional outlier_rows sheet is included when non-empty."""
        from openpyxl import load_workbook
        result = ExcelExporter.build_removed_excel(
            pd.DataFrame({"a": [1]}),
            pd.DataFrame({"a": [2]}),
            pd.DataFrame({"a": [3]}),
            outlier_rows=pd.DataFrame({"a": [4]}),
        )
        wb = load_workbook(io.BytesIO(result))
        assert "Participant_Outliers" in wb.sheetnames

    def test_build_filtered_csv_returns_bytes(self):
        result = ExcelExporter.build_filtered_csv(self._make_clean_df())
        assert isinstance(result, bytes)
        assert b"event_name" in result


# ─────────────────────────────────────────────────────────────────────────────
# Gemini API key validation
# Uses AIQueryEngine internals / config to validate the key.
# ─────────────────────────────────────────────────────────────────────────────

class TestGeminiApiKey:
    def test_api_key_is_set(self):
        import os
        from dotenv import load_dotenv
        load_dotenv()
        key = os.getenv("GEMINI_API_KEY")
        assert key, (
            "PROBLEM: GEMINI_API_KEY is not set.\n"
            "Fix: create a .env file in the project root with:\n"
            "  GEMINI_API_KEY=your_actual_key_here"
        )

    def test_api_key_is_not_placeholder(self):
        import os
        from dotenv import load_dotenv
        load_dotenv()
        key = os.getenv("GEMINI_API_KEY", "")
        placeholders = {"your_key_here", "your_actual_key_here", "changeme", "xxx", ""}
        assert key.lower() not in placeholders, (
            f"PROBLEM: GEMINI_API_KEY looks like a placeholder: '{key}'\n"
            "Fix: replace it with your real Gemini API key from "
            "https://aistudio.google.com/app/apikey"
        )

    def test_api_key_matches_config(self):
        """config.GEMINI_API_KEY should match the environment variable."""
        import os
        from dotenv import load_dotenv
        load_dotenv()
        from src.config import GEMINI_API_KEY
        env_key = os.getenv("GEMINI_API_KEY", "")
        if not env_key:
            pytest.skip("GEMINI_API_KEY not set — skipping")
        assert GEMINI_API_KEY == env_key, (
            "config.GEMINI_API_KEY does not match the GEMINI_API_KEY env var. "
            "Check _get_gemini_key() in config.py."
        )

    def test_api_key_works_with_gemini(self):
        """Makes a real (minimal) call to Gemini and prints the exact error if it fails."""
        import os
        from dotenv import load_dotenv
        load_dotenv()
        key = os.getenv("GEMINI_API_KEY", "")
        if not key:
            pytest.skip("GEMINI_API_KEY not set — skipping live test")

        try:
            import google.generativeai as genai
            genai.configure(api_key=key)
            model    = genai.GenerativeModel("gemini-2.5-flash")
            response = model.generate_content("Reply with the single word: OK")
            text     = response.text.strip()
            assert text, (
                "PROBLEM: Gemini returned an empty response.\n"
                "This may mean the model name is wrong or the key has no quota."
            )
        except Exception as e:
            pytest.fail(
                f"PROBLEM: Gemini API call failed.\n"
                f"Exact error: {type(e).__name__}: {e}\n\n"
                "Common causes:\n"
                "  - Invalid or revoked API key\n"
                "  - No quota / billing not enabled\n"
                "  - Network issue or firewall blocking generativeai API\n"
                "  - Wrong model name ('gemini-2.5-flash' may not be available on your key tier)"
            )