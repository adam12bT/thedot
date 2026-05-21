

import io
import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelExporter:
    """Static methods for building Excel / CSV download artefacts."""

    # ── Cleaned events workbook ────────────────────────────────────────────────

    @staticmethod
    def build_cleaned_excel(df: pd.DataFrame) -> bytes:
        """Return a styled .xlsx workbook with an EVENTS sheet and a SUMMARY sheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "EVENTS"

        # Style objects
        hf  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hfi = PatternFill("solid", fgColor="4F46E5")
        ha  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cf  = Font(name="Calibri", size=10)
        ca  = Alignment(vertical="center")
        alt = PatternFill("solid", fgColor="F8F9FF")
        tb  = Border(
            left=Side(style="thin", color="EBEBEB"),
            right=Side(style="thin", color="EBEBEB"),
            bottom=Side(style="thin", color="EBEBEB"),
        )

        export_cols = [
            "event_name", "activity_type", "start_time", "end_time", "room",
            "booking_date", "organizer_email", "organizer_name", "organization",
            "participants", "status", "signed_declaration", "comment",
        ]
        labels = [
            "Event Name", "Activity Type", "Start Time", "End Time", "Room",
            "Booking Date", "Organizer Email", "Organizer Name", "Organization",
            "Participants", "Status", "Signed Declaration", "Comment",
        ]

        # Header row
        for ci, lbl in enumerate(labels, 1):
            c = ws.cell(row=1, column=ci, value=lbl)
            c.font = hf; c.fill = hfi; c.alignment = ha
        ws.row_dimensions[1].height = 28

        # Data rows
        for ri, (_, row_data) in enumerate(df[export_cols].iterrows(), 2):
            fill = alt if ri % 2 == 0 else None
            for ci, col in enumerate(export_cols, 1):
                val = row_data[col]
                c = ws.cell(
                    row=ri, column=ci,
                    value=None if (not isinstance(val, str) and pd.isna(val)) else val,
                )
                c.font = cf; c.alignment = ca; c.border = tb
                if fill:
                    c.fill = fill
                if col in ("start_time", "end_time") and isinstance(val, datetime.datetime):
                    c.number_format = "DD/MM/YYYY HH:MM"
                elif col == "booking_date" and isinstance(val, datetime.datetime):
                    c.number_format = "DD/MM/YYYY"

        # Column widths
        for ci, w in enumerate([32, 22, 18, 18, 22, 15, 32, 24, 26, 13, 12, 18, 22], 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

        # Data-validation dropdowns
        last_row = len(df) + 1
        for dv_type, formula, sqref in [
            ("list", '"Evénement interne,Evénement externe,Visite officielle,Visite estudantine,Evénement SSO/ Startup"', f"B2:B{last_row}"),
            ("list", '"Salle de formation,Salle Fondation,Think room,Salle de réunion 113,Terrasse"',                    f"E2:E{last_row}"),
            ("list", '"Tenu,Annulé,Reporté"',                                                                            f"K2:K{last_row}"),
        ]:
            dv = DataValidation(type=dv_type, formula1=formula, allow_blank=True, showDropDown=False)
            dv.sqref = sqref
            ws.add_data_validation(dv)

        # Table
        tbl = Table(displayName="Evenements", ref=f"A1:M{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
        ws.freeze_panes = "A2"

        # SUMMARY sheet
        ExcelExporter._add_summary_sheet(wb, df)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _add_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
        ws2 = wb.create_sheet("SUMMARY")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 22
        ws2.cell(row=1, column=1, value="Summary").font = Font(bold=True, size=13, color="4F46E5")

        total = len(df)
        held  = int((df["status"] == "Tenu").sum())
        canc  = int((df["status"] == "Annulé").sum())

        rows = [
            ("Total Events",          total),
            ("Held Events",           held),
            ("Cancelled Events",      canc),
            ("Cancellation Rate (%)", round(canc / total * 100, 1) if total else 0),
            ("Total Hours Booked",    round(df["duration_hours"].sum(), 1)),
            ("Total Participants",    int(df["participants"].sum(skipna=True))),
            ("Avg Participants",      round(df["participants"].mean(skipna=True), 1)),
            ("Unique Organizations",  int(df["organization"].nunique())),
            ("Unique Rooms",          int(df["room"].nunique())),
        ]
        for ri, (lbl, val) in enumerate(rows, 2):
            ws2.cell(row=ri, column=1, value=lbl).font = Font(bold=True, name="Calibri", size=10)
            ws2.cell(row=ri, column=2, value=val).font  = Font(name="Calibri", size=10)

    # ── Removed rows workbook ──────────────────────────────────────────────────

    @staticmethod
    def build_removed_excel(
        empty_rows: pd.DataFrame,
        duplicate_rows: pd.DataFrame,
        negative_rows: pd.DataFrame,
        outlier_rows: pd.DataFrame | None = None,
    ) -> bytes:
        """Return an .xlsx workbook with one sheet per removal reason."""
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            empty_rows.to_excel(writer, sheet_name="Empty", index=False)
            duplicate_rows.to_excel(writer, sheet_name="Duplicates", index=False)
            negative_rows.to_excel(writer, sheet_name="Negative_Duration", index=False)
            if outlier_rows is not None and not outlier_rows.empty:
                outlier_rows.to_excel(writer, sheet_name="Participant_Outliers", index=False)
        return buf.getvalue()

    # ── CSV ────────────────────────────────────────────────────────────────────

    @staticmethod
    def build_filtered_csv(df: pd.DataFrame) -> bytes:
        """Return the current filtered DataFrame as UTF-8 CSV bytes."""
        return df.to_csv(index=False).encode()
