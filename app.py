import re
import datetime
import io
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.patches import Patch
import seaborn as sns
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import cross_val_score, StratifiedKFold
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import confusion_matrix
from sklearn.pipeline import Pipeline
import json
import google.generativeai as genai
from dotenv import load_dotenv

load_dotenv()

import os
_GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
genai.configure(api_key=_GEMINI_API_KEY)

GEMINI_MODEL = "gemini-2.5-flash"


# ══════════════════════════════════════════════════════════════════════════════
# CHAT / AI QUERY
# ══════════════════════════════════════════════════════════════════════════════

def _auto_answer(question: str, result_df) -> str | None:
    """Generate a factual answer from result_df when possible, overriding LLM hallucination."""
    if result_df is None or result_df.empty:
        return None
    if result_df.shape == (1, 1):
        val = result_df.iloc[0, 0]
        col = result_df.columns[0]
        if col == "result":
            if isinstance(val, (int, np.integer)):
                return f"The answer is **{int(val):,}**."
            elif isinstance(val, (float, np.floating)):
                return f"The answer is **{float(val):,.2f}**."
            else:
                return f"The answer is **{val}**."
    return None


def _fix_answer_number(answer: str, result_df) -> str:
    """If result_df is a single scalar and the answer states a wrong number, correct it."""
    if result_df is None or result_df.empty:
        return answer
    try:
        if result_df.shape == (1, 1):
            actual_value = result_df.iloc[0, 0]
            if not isinstance(actual_value, (int, float, np.integer, np.floating)):
                return answer
            actual = int(actual_value) if isinstance(actual_value, (int, np.integer)) else float(actual_value)
            nums_in_answer = re.findall(r'\b\d+(?:\.\d+)?\b', answer)
            if nums_in_answer:
                stated_str = nums_in_answer[0]
                stated = float(stated_str)
                if stated != actual:
                    actual_str = str(int(actual)) if isinstance(actual, float) and actual == int(actual) else str(actual)
                    answer = re.sub(r'\b' + re.escape(stated_str) + r'\b', actual_str, answer, count=1)
    except Exception:
        pass
    return answer


def text_to_pandas(question: str, df, history: list) -> dict:

    if not _GEMINI_API_KEY:
        return {
            "answer": "", "thought": "", "code": "",
            "result_df": None,
            "error": "No Gemini API key set. Add GEMINI_API_KEY to your environment variables.",
        }

    system_prompt = f"""You are a Python/pandas expert assistant.
The user has a DataFrame called `df` with the following structure:
- Columns : {list(df.columns)}
- Dtypes  : {df.dtypes.to_dict()}
- Sample  : {df.head(3).to_dict()}

Rules:
1. Write pandas code that answers the question using the variable `df`.
2. Store the final result in a variable called `result_df` (always a DataFrame or None).
3. If the result is a single number or value, wrap it: result_df = pd.DataFrame([{{"result": <value>}}])
4. Return ONLY a valid JSON object — no markdown fences, no extra text — with these exact keys:
   - "thought" : your brief reasoning (1-2 sentences)
   - "code"    : the executable Python/pandas code. IMPORTANT: use only single quotes for Python strings inside the code field to avoid JSON parse errors. Never use double quotes inside Python string literals in the code.
   - "answer"  : plain-English answer. CRITICAL: if your code computes a scalar result stored in result_df, you MUST read that computed value and use it in your answer. Never guess or assume the number — always derive the answer from what your code actually computes.

CRITICAL JSON RULES:
- The entire response must be valid JSON.
- Never use unescaped double quotes inside string values.
- Never include raw newline characters inside string values — use \\n if needed.
- Use single quotes for all Python string literals in the code field.

Example output format:
{{"thought":"I will filter for cancelled events in the last year and count them.","code":"last_year = df['year'].max() - 1\\nfiltered = df[(df['year'] == last_year) & (df['status'] == 'Annulé')]\\nresult_df = pd.DataFrame([{{'result': len(filtered)}}])","answer":"There were 42 events cancelled last year."}}"""

    gemini_history = []
    for msg in history:
        role = "model" if msg["role"] == "assistant" else "user"
        gemini_history.append({"role": role, "parts": [msg["content"]]})

    raw = ""
    try:
        model = genai.GenerativeModel(
            model_name=GEMINI_MODEL,
            system_instruction=system_prompt,
            generation_config=genai.GenerationConfig(temperature=0.1),
        )
        chat = model.start_chat(history=gemini_history)
        response = chat.send_message(question)
        raw = response.text.strip()

        # Step 1: strip markdown fences
        clean = re.sub(r"```(?:json|python)?|```", "", raw).strip()

        # Step 2: extract the JSON object
        match = re.search(r"\{.*\}", clean, re.DOTALL)
        if not match:
            return {
                "answer": "", "thought": "", "code": raw,
                "result_df": None, "error": "Model did not return valid JSON.",
            }

        json_str = match.group()

        # Step 3: try direct parse
        parsed = None
        try:
            parsed = json.loads(json_str)
        except json.JSONDecodeError:
            # Step 4: fix unescaped newlines and tabs inside string values
            json_str_fixed = re.sub(r'(?<!\\)\n', r'\\n', json_str)
            json_str_fixed = re.sub(r'(?<!\\)\t', r'\\t', json_str_fixed)
            try:
                parsed = json.loads(json_str_fixed)
            except json.JSONDecodeError:
                # Step 5: last resort — regex field extraction
                thought_m = re.search(r'"thought"\s*:\s*"((?:[^"\\]|\\.)*)"', json_str)
                answer_m  = re.search(r'"answer"\s*:\s*"((?:[^"\\]|\\.)*)"',  json_str)
                # For code, try to grab everything between "code": " and the next top-level key
                code_m    = re.search(r'"code"\s*:\s*"(.*?)"(?=\s*,\s*"(?:answer|thought)")', json_str, re.DOTALL)
                if not code_m:
                    code_m = re.search(r'"code"\s*:\s*"((?:[^"\\]|\\.)*)"', json_str)

                parsed = {
                    "thought": thought_m.group(1) if thought_m else "",
                    "code":    code_m.group(1).replace("\\n", "\n") if code_m else "",
                    "answer":  answer_m.group(1) if answer_m else "",
                }

        if parsed is None:
            return {
                "answer": "", "thought": "", "code": raw,
                "result_df": None, "error": "JSON parse error: could not recover from malformed response.",
            }

        code = parsed.get("code", "").strip()
        # Unescape sequences Gemini may have double-escaped
        code = code.replace("\\n", "\n").replace("\\'", "'")

        result_df = None
        if code:
            local_vars = {"df": df.copy(), "pd": pd, "np": np}
            exec(code, {}, local_vars)  # nosec
            result_df = local_vars.get("result_df", None)

            if isinstance(result_df, pd.Series):
                result_df = result_df.to_frame()
            elif isinstance(result_df, (int, float, str, bool, np.integer, np.floating)):
                result_df = pd.DataFrame([{"result": result_df}])
            elif isinstance(result_df, (list, tuple, dict, np.ndarray)):
                try:
                    result_df = pd.DataFrame(result_df)
                except Exception:
                    result_df = pd.DataFrame([{"result": str(result_df)}])
            elif result_df is not None and not isinstance(result_df, pd.DataFrame):
                try:
                    result_df = pd.DataFrame([{"result": str(result_df)}])
                except Exception:
                    result_df = None

        # ── Fix answer: prefer auto-generated truth over LLM's stated answer ──
        answer_text = parsed.get("answer", "")
        auto = _auto_answer(question, result_df)
        if auto:
            answer_text = auto
        else:
            answer_text = _fix_answer_number(answer_text, result_df)

        return {
            "answer":    answer_text,
            "thought":   parsed.get("thought", ""),
            "code":      code,
            "result_df": result_df,
            "error":     None,
        }

    except json.JSONDecodeError as e:
        return {
            "answer": "", "thought": "", "code": raw,
            "result_df": None, "error": f"JSON parse error: {e}",
        }
    except Exception as e:
        return {
            "answer": "", "thought": "", "code": raw,
            "result_df": None, "error": str(e),
        }


# ══════════════════════════════════════════════════════════════════════════════
# PALETTE
# ══════════════════════════════════════════════════════════════════════════════
INDIGO  = "#4F46E5"
TEAL    = "#0D9488"
ROSE    = "#E11D48"
AMBER   = "#D97706"
EMERALD = "#059669"
VIOLET  = "#7C3AED"
SKY     = "#0284C7"
BLACK   = "#1A1A2E"
WHITE   = "#FFFFFF"
MGRAY   = "#94A3B8"
LGRAY   = "#F8FAFC"

CHART_COLORS = [
    "#4F46E5", "#0D9488", "#E11D48", "#D97706",
    "#059669", "#7C3AED", "#0284C7", "#DB2777",
    "#065F46", "#92400E",
]

PARTICIPANT_MAX = 1000


def remove_outliers(series):
    return series[(series >= 0) & (series <= PARTICIPANT_MAX)]


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_date(s):
    if pd.isna(s):
        return None
    s = str(s).strip()
    if s == '':
        return None
    m = re.search(r'(\w{3} \w{3} \d{2} \d{4} \d{2}:\d{2}:\d{2})', s)
    if m:
        try:
            return pd.to_datetime(m.group(1), format='%a %b %d %Y %H:%M:%S')
        except Exception:
            pass
    m = re.search(r'(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})', s)
    if m:
        try:
            return pd.to_datetime(m.group(1))
        except Exception:
            pass
    try:
        return pd.to_datetime(s)
    except Exception:
        return None


def find_col(df, *candidates):
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


def map_status(s):
    if pd.isna(s):
        return s
    return {
        'finished': 'Tenu', 'canceled': 'Annulé', 'rejected': 'Annulé',
        'accepted': 'Tenu', 'tenu': 'Tenu', 'annulé': 'Annulé', 'annule': 'Annulé',
    }.get(str(s).strip().lower(), str(s).strip())


def map_space(s):
    if pd.isna(s) or str(s).strip() == '':
        return ''
    return {
        'Salle Polyvalente': 'Salle de formation',
        'ESPACE FONDATION ': 'Salle Fondation',
        'Training Room': 'Salle de formation',
        'Terrasse': 'Terrasse',
        'BALE': 'BALE',
        'Podcast Tunisia': 'Podcast',
        'Salle Design Thinking': 'Salle Design Thinking',
        'Salle de réunion 113': 'Salle de réunion 113',
    }.get(str(s).strip(), str(s).strip())


# ══════════════════════════════════════════════════════════════════════════════
# CHART STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def chart_style():
    plt.rcParams.update({
        'font.family': 'DejaVu Sans',
        'axes.spines.top': False, 'axes.spines.right': False,
        'axes.spines.left': False, 'axes.spines.bottom': True,
        'axes.edgecolor': '#EBEBEB', 'axes.facecolor': WHITE,
        'figure.facecolor': WHITE, 'grid.color': '#F4F4F8',
        'grid.linewidth': 0.6, 'xtick.color': MGRAY, 'ytick.color': MGRAY,
        'axes.titlesize': 12, 'axes.titleweight': '600',
        'axes.titlecolor': BLACK, 'axes.titlepad': 14,
    })


def make_fig(w=9, h=4.5):
    fig, ax = plt.subplots(figsize=(w, h), facecolor=WHITE)
    ax.set_facecolor(WHITE)
    ax.spines[['top', 'right', 'left']].set_visible(False)
    ax.spines['bottom'].set_color('#EBEBEB')
    ax.tick_params(colors=MGRAY, length=0)
    ax.yaxis.grid(True, color='#F4F4F8', linewidth=0.7)
    ax.set_axisbelow(True)
    return fig, ax


# ══════════════════════════════════════════════════════════════════════════════
# DATA PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def load_and_clean(file_bytes, file_name):
    ext = Path(file_name).suffix.lower()
    df_raw = (
        pd.read_csv(io.BytesIO(file_bytes), dtype=str)
        if ext == '.csv'
        else pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    )

    COL_TITLE        = find_col(df_raw, 'title', 'event_name', 'name', 'nom')
    COL_START        = find_col(df_raw, 'startTime', 'start_time', 'start', 'début')
    COL_END          = find_col(df_raw, 'endTime', 'end_time', 'end', 'fin')
    COL_STATUS       = find_col(df_raw, 'status', 'statut', 'état')
    COL_VISIBILITY   = find_col(df_raw, 'visibility', 'type', 'visibilité')
    COL_SPACE        = find_col(df_raw, 'event_proposals.space.name', 'room', 'salle', 'space')
    COL_EMAIL        = find_col(df_raw, 'organizer.email', 'email', 'courriel')
    COL_FIRSTNAME    = find_col(df_raw, 'organizer.firstName', 'firstName', 'prenom', 'first_name')
    COL_LASTNAME     = find_col(df_raw, 'organizer.lastname', 'lastName', 'last_name')
    COL_ORG          = find_col(df_raw, 'organizer.organization.name', 'organization', 'organisation', 'org')
    COL_PARTICIPANTS = find_col(df_raw, 'participantNb', 'participants', 'participant_count')
    COL_POLICY       = find_col(df_raw, 'policy', 'declaration', 'signed')
    COL_THEME        = find_col(df_raw, 'theme', 'thème', 'category', 'catégorie')
    COL_BOOKING      = find_col(df_raw, 'booking_date', 'bookingDate', 'reservation_date')

    def gv(row, col):
        return row[col] if col and col in row.index else None

    records = []
    for _, row in df_raw.iterrows():
        fn = gv(row, COL_FIRSTNAME) or ''
        ln = gv(row, COL_LASTNAME) or ''
        vis = gv(row, COL_VISIBILITY)
        activity_type = (
            'Evénement externe' if vis and str(vis).lower() == 'public'
            else ('Evénement interne' if vis else None)
        )
        policy_val = gv(row, COL_POLICY)
        records.append({
            'event_name':         gv(row, COL_TITLE),
            'activity_type':      activity_type,
            'start_time':         parse_date(gv(row, COL_START)),
            'end_time':           parse_date(gv(row, COL_END)),
            'room':               map_space(gv(row, COL_SPACE)),
            'booking_date':       parse_date(gv(row, COL_BOOKING)),
            'organizer_email':    gv(row, COL_EMAIL),
            'organizer_name':     f'{fn} {ln}'.strip(),
            'organization':       gv(row, COL_ORG),
            'participants':       pd.to_numeric(gv(row, COL_PARTICIPANTS), errors='coerce'),
            'status':             map_status(gv(row, COL_STATUS)),
            'signed_declaration': (
                'Oui' if pd.notna(policy_val) and str(policy_val).strip()
                not in ['', 'nan', 'NaN', '-'] else 'Non'
            ),
            'comment':            gv(row, COL_THEME),
        })

    df = pd.DataFrame(records)
    key_cols = [c for c in ['event_name', 'start_time', 'end_time', 'status'] if c in df.columns]
    empty_rows = df[df[key_cols].isna().all(axis=1)].copy()
    df = df.dropna(subset=key_cols, how='all').reset_index(drop=True)

    dupe_cols = [c for c in ['event_name', 'start_time', 'end_time', 'room'] if c in df.columns]
    dupe_mask = df.duplicated(subset=dupe_cols, keep='first')
    duplicate_rows = df[dupe_mask].copy()
    df = df[~dupe_mask].reset_index(drop=True)

    activity_map = {
        'événement interne': 'Evénement interne',
        'evénement interne': 'Evénement interne',
        'evenement interne': 'Evénement interne',
        'evénement externe': 'Evénement externe',
        'evenement externe': 'Evénement externe',
    }
    df['activity_type'] = df['activity_type'].apply(
        lambda x: activity_map.get(x.lower() if x else x, x) if pd.notna(x) else x
    )
    df['status'] = df['status'].str.strip().replace({'Reporté': 'Annulé'})

    def fix_time(dt):
        if pd.isna(dt):
            return dt
        if isinstance(dt, datetime.datetime) and dt.hour == 0 and dt.minute < 24:
            return dt.replace(hour=dt.minute, minute=0)
        return dt

    df['start_time'] = df['start_time'].apply(fix_time)
    df['end_time']   = df['end_time'].apply(fix_time)
    df['duration_hours'] = (
        (df['end_time'] - df['start_time']).dt.total_seconds() / 3600
    ).round(2)

    neg_mask = df['duration_hours'] < 0
    negative_rows = df[neg_mask].copy()
    df.loc[neg_mask, ['start_time', 'end_time']] = df.loc[neg_mask, ['end_time', 'start_time']].values
    df.loc[neg_mask, 'duration_hours'] = df.loc[neg_mask, 'duration_hours'].abs()

    df['duration_flag'] = df['duration_hours'].apply(
        lambda x: 'multi-day' if pd.notna(x) and x > 24 else ''
    )
    df['date']       = df['start_time'].dt.date
    df['year']       = df['start_time'].dt.year
    df['month']      = df['start_time'].dt.month
    df['month_name'] = df['start_time'].dt.strftime('%b')
    df['weekday']    = df['start_time'].dt.day_name()
    df['hour']       = df['start_time'].dt.hour
    df['room']       = df['room'].str.replace(r'^0,\s*', '', regex=True).str.strip()

    df['participant_outlier'] = ~df['participants'].isna() & (
        (df['participants'] < 0) | (df['participants'] > PARTICIPANT_MAX)
    )

    return df, empty_rows, duplicate_rows, negative_rows


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = 'EVENTS'

    hf  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hfi = PatternFill('solid', fgColor='4F46E5')
    ha  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cf  = Font(name='Calibri', size=10)
    ca  = Alignment(vertical='center')
    alt = PatternFill('solid', fgColor='F8F9FF')
    tb  = Border(
        left=Side(style='thin', color='EBEBEB'),
        right=Side(style='thin', color='EBEBEB'),
        bottom=Side(style='thin', color='EBEBEB'),
    )

    export_cols = [
        'event_name', 'activity_type', 'start_time', 'end_time', 'room', 'booking_date',
        'organizer_email', 'organizer_name', 'organization', 'participants',
        'status', 'signed_declaration', 'comment',
    ]
    labels = [
        'Event Name', 'Activity Type', 'Start Time', 'End Time', 'Room', 'Booking Date',
        'Organizer Email', 'Organizer Name', 'Organization', 'Participants',
        'Status', 'Signed Declaration', 'Comment',
    ]

    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=1, column=ci, value=lbl)
        c.font = hf; c.fill = hfi; c.alignment = ha
    ws.row_dimensions[1].height = 28

    for ri, (_, row_data) in enumerate(df[export_cols].iterrows(), 2):
        fill = alt if ri % 2 == 0 else None
        for ci, col in enumerate(export_cols, 1):
            val = row_data[col]
            c = ws.cell(
                row=ri, column=ci,
                value=None if (not isinstance(val, str) and pd.isna(val)) else val,
            )
            c.font = cf; c.alignment = ca; c.border = tb
            if fill:
                c.fill = fill
            if col in ('start_time', 'end_time') and isinstance(val, datetime.datetime):
                c.number_format = 'DD/MM/YYYY HH:MM'
            elif col == 'booking_date' and isinstance(val, datetime.datetime):
                c.number_format = 'DD/MM/YYYY'

    for ci, w in enumerate([32, 22, 18, 18, 22, 15, 32, 24, 26, 13, 12, 18, 22], 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    last_row = len(df) + 1
    for spec in [
        ('list', '"Evénement interne,Evénement externe,Visite officielle,Visite estudantine,Evénement SSO/ Startup"', f'B2:B{last_row}'),
        ('list', '"Salle de formation,Salle Fondation,Think room,Salle de réunion 113,Terrasse"',                    f'E2:E{last_row}'),
        ('list', '"Tenu,Annulé,Reporté"',                                                                            f'K2:K{last_row}'),
    ]:
        dv = DataValidation(type=spec[0], formula1=spec[1], allow_blank=True, showDropDown=False)
        dv.sqref = spec[2]
        ws.add_data_validation(dv)

    tbl = Table(displayName='Evenements', ref=f'A1:M{last_row}')
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tbl)
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet('SUMMARY')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 22
    ws2.cell(row=1, column=1, value='Summary').font = Font(bold=True, size=13, color='4F46E5')

    total = len(df)
    held  = int((df['status'] == 'Tenu').sum())
    canc  = int((df['status'] == 'Annulé').sum())
    clean_p = remove_outliers(df['participants'])

    for ri, (lbl, val) in enumerate([
        ('Total Events',                    total),
        ('Held Events',                     held),
        ('Cancelled Events',                canc),
        ('Cancellation Rate (%)',           round(canc / total * 100, 1) if total else 0),
        ('Total Hours Booked',              round(df['duration_hours'].sum(), 1)),
        ('Total Participants (excl. outliers)', int(clean_p.sum(skipna=True))),
        ('Avg Participants (excl. outliers)',   round(clean_p.mean(skipna=True), 1)),
        ('Unique Organizations',            int(df['organization'].nunique())),
        ('Unique Rooms',                    int(df['room'].nunique())),
    ], 2):
        ws2.cell(row=ri, column=1, value=lbl).font = Font(bold=True, name='Calibri', size=10)
        ws2.cell(row=ri, column=2, value=val).font = Font(name='Calibri', size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_removed_excel(empty_rows, duplicate_rows, negative_rows):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        empty_rows.to_excel(w, sheet_name='Empty', index=False)
        duplicate_rows.to_excel(w, sheet_name='Duplicates', index=False)
        negative_rows.to_excel(w, sheet_name='Negative_Duration', index=False)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# CHART FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def c_status(df):
    counts = df['status'].value_counts()
    fig, ax = make_fig(7, 4)
    color_map = {'Tenu': EMERALD, 'Annulé': ROSE}
    colors = [color_map.get(s, MGRAY) for s in counts.index]
    bars = ax.bar(counts.index, counts.values, color=colors, edgecolor=WHITE,
                  linewidth=2, width=0.42, alpha=0.92)
    for b in bars:
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.4,
                f"{int(b.get_height()):,}", ha='center', fontweight='600',
                fontsize=12, color=BLACK)
    ax.set_title('Events by Status')
    ax.set_ylabel('Count', color=MGRAY, fontsize=10)
    ax.tick_params(axis='x', rotation=0, labelsize=11)
    plt.tight_layout()
    return fig


def c_activity(df):
    counts = df['activity_type'].value_counts().dropna()
    fig, ax = make_fig(9, max(3.5, len(counts) * 0.65 + 1.5))
    colors = CHART_COLORS[:len(counts)]
    ax.barh(counts.index, counts.values, color=colors, edgecolor=WHITE, height=0.52, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.3, p.get_y() + p.get_height() / 2,
                f"{int(p.get_width()):,}", va='center', fontsize=10, color=MGRAY, fontweight='500')
    ax.set_title('Events by Activity Type')
    ax.set_xlabel('Count', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


def c_monthly(df):
    monthly = df.groupby(['year', 'month']).size().reset_index(name='count')
    monthly['period'] = pd.to_datetime(monthly[['year', 'month']].assign(day=1))
    monthly = monthly.sort_values('period')
    fig, ax = make_fig(max(9, len(monthly) * 0.8 + 2), 4.5)
    ax.fill_between(monthly['period'], monthly['count'], alpha=0.08, color=INDIGO)
    ax.plot(monthly['period'], monthly['count'], marker='o', color=INDIGO,
            linewidth=2.5, markersize=6, zorder=3, markerfacecolor=WHITE, markeredgewidth=2.5)
    for x, y in zip(monthly['period'], monthly['count']):
        ax.annotate(str(y), (x, y), textcoords='offset points', xytext=(0, 9),
                    ha='center', fontsize=8, color=MGRAY, fontweight='500')
    ax.set_title('Events per Month')
    ax.set_ylabel('Events', color=MGRAY, fontsize=10)
    ax.xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%b %Y'))
    plt.xticks(rotation=40, ha='right', fontsize=9)
    plt.tight_layout()
    return fig


def c_weekday(df):
    order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    counts = df['weekday'].value_counts().reindex(order, fill_value=0)
    fig, ax = make_fig(9, 4.5)
    max_v = counts.max()
    colors = [INDIGO if v == max_v and max_v > 0 else '#CBD5E1' for v in counts.values]
    bars = ax.bar(counts.index, counts.values, color=colors, edgecolor=WHITE, width=0.6, alpha=0.92)
    for b in bars:
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.3,
                f"{int(b.get_height()):,}", ha='center', fontsize=9, color=MGRAY, fontweight='500')
    ax.set_title('Events by Day of Week')
    ax.set_ylabel('Count', color=MGRAY, fontsize=10)
    ax.tick_params(axis='x', rotation=15)
    plt.tight_layout()
    return fig


def c_hour(df):
    hc = df['hour'].value_counts().sort_index()
    all_h = pd.Series(0, index=range(24))
    all_h.update(hc)
    fig, ax = make_fig(11, 4.2)
    slot_colors = {
        range(0, 6): '#7C3AED', range(6, 9): '#D97706',
        range(9, 12): '#0284C7', range(12, 14): '#D97706',
        range(14, 18): '#059669', range(18, 21): '#E11D48',
        range(21, 24): '#0D9488',
    }
    hour_palette = []
    for h in range(24):
        for rng, c in slot_colors.items():
            if h in rng:
                hour_palette.append(c)
                break
    max_v = all_h.max()
    colors = [
        INDIGO if v == max_v and max_v > 0 else hour_palette[i]
        for i, v in enumerate(all_h.values)
    ]
    ax.bar(all_h.index, all_h.values, color=colors, edgecolor=WHITE, width=0.75, alpha=0.88)
    ax.set_title('Events by Start Hour')
    ax.set_xlabel('Hour of Day', color=MGRAY, fontsize=10)
    ax.set_ylabel('Count', color=MGRAY, fontsize=10)
    ax.set_xticks(range(24))
    legend_items = [
        Patch(color='#7C3AED', label='Night (0–5)'),
        Patch(color='#D97706', label='Early AM (6–8)'),
        Patch(color='#0284C7', label='Morning (9–11)'),
        Patch(color='#D97706', label='Noon (12–13)'),
        Patch(color='#059669', label='Afternoon (14–17)'),
        Patch(color='#E11D48', label='Evening (18–20)'),
        Patch(color='#0D9488', label='Late (21–23)'),
    ]
    ax.legend(handles=legend_items, loc='upper right', fontsize=7,
              framealpha=0.95, edgecolor='#EBEBEB', ncol=2)
    plt.tight_layout()
    return fig


def c_top_rooms(df):
    top = df['room'].replace('', pd.NA).dropna().value_counts().head(10)
    fig, ax = make_fig(9, max(3.5, len(top) * 0.6 + 1))
    colors = CHART_COLORS[:len(top)]
    ax.barh(top.index, top.values, color=colors, edgecolor=WHITE, height=0.58, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.2, p.get_y() + p.get_height() / 2,
                f"{int(p.get_width()):,}", va='center', fontsize=9, color=MGRAY, fontweight='500')
    ax.set_title('Top 10 Most Used Rooms')
    ax.set_xlabel('Events', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


def c_top_orgs(df):
    top = df['organization'].replace('', pd.NA).dropna().value_counts().head(15)
    fig, ax = make_fig(9, max(4, len(top) * 0.5 + 1))
    colors = [INDIGO] + CHART_COLORS[1:len(top)]
    ax.barh(top.index, top.values, color=colors[:len(top)], edgecolor=WHITE, height=0.58, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.1, p.get_y() + p.get_height() / 2,
                f"{int(p.get_width()):,}", va='center', fontsize=8, color=MGRAY, fontweight='500')
    ax.set_title('Top 15 Organizations')
    ax.set_xlabel('Events', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


def c_avg_participants(df):
    clean = df[~df['participant_outlier']].copy()
    avg = clean.groupby('activity_type')['participants'].mean().dropna().sort_values(ascending=False)
    fig, ax = make_fig(9, max(3.5, len(avg) * 0.65 + 1))
    colors = CHART_COLORS[:len(avg)]
    ax.barh(avg.index, avg.values, color=colors, edgecolor=WHITE, height=0.52, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.5, p.get_y() + p.get_height() / 2,
                f'{p.get_width():.1f}', va='center', fontsize=9, color=MGRAY, fontweight='500')
    ax.set_title('Avg Participants by Activity Type (outliers excluded)')
    ax.set_xlabel('Avg Participants', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


def c_cancel_rate(df):
    rs = df.groupby(['room', 'status']).size().unstack(fill_value=0)
    if 'Annulé' not in rs.columns:
        return None
    rs['rate'] = (rs['Annulé'] / rs.sum(axis=1) * 100).round(1)
    top = rs['rate'].sort_values(ascending=False).head(10)
    top = top[top > 0]
    if top.empty:
        return None
    fig, ax = make_fig(9, max(3.5, len(top) * 0.65 + 1))
    max_r = top.max()
    colors = [ROSE if v == max_r else '#FDA4AF' for v in top.values]
    ax.barh(top.index, top.values, color=colors, edgecolor=WHITE, height=0.58, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.3, p.get_y() + p.get_height() / 2,
                f'{p.get_width():.1f}%', va='center', fontsize=9, color=MGRAY, fontweight='500')
    ax.xaxis.set_major_formatter(mticker.PercentFormatter())
    ax.set_title('Cancellation Rate by Room')
    ax.set_xlabel('Cancellation Rate', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


def c_room_hours(df):
    rd = df.groupby('room')['duration_hours'].sum().sort_values(ascending=False).head(10)
    rd = rd[rd.index != '']
    fig, ax = make_fig(9, max(3.5, len(rd) * 0.65 + 1))
    colors = CHART_COLORS[:len(rd)]
    ax.barh(rd.index, rd.values, color=colors, edgecolor=WHITE, height=0.58, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.5, p.get_y() + p.get_height() / 2,
                f'{p.get_width():.0f}h', va='center', fontsize=9, color=MGRAY, fontweight='500')
    ax.set_title('Rooms by Total Reservation Hours')
    ax.set_xlabel('Total Hours', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


def c_participants_dist(df):
    all_data = df['participants'].dropna()
    clean    = remove_outliers(all_data)
    n_out    = len(all_data) - len(clean)

    fig, axes = plt.subplots(1, 2, figsize=(13, 4.5), facecolor=WHITE)
    for ax in axes:
        ax.set_facecolor(WHITE)
        ax.spines[['top', 'right', 'left']].set_visible(False)
        ax.spines['bottom'].set_color('#EBEBEB')
        ax.tick_params(colors=MGRAY, length=0)
        ax.yaxis.grid(True, color='#F4F4F8', linewidth=0.7)
        ax.set_axisbelow(True)

    axes[0].boxplot(
        clean, vert=True, patch_artist=True,
        boxprops=dict(facecolor='#EEF2FF', alpha=0.9, linewidth=1.5),
        medianprops=dict(color=INDIGO, linewidth=2.5),
        whiskerprops=dict(color=TEAL, linewidth=1.5),
        capprops=dict(color=TEAL, linewidth=2),
        flierprops=dict(marker='o', color=ROSE, alpha=0.5, markersize=4),
    )
    axes[0].set_title('Participant Distribution\n(outliers removed)', fontsize=11)
    axes[0].set_ylabel('Participants', color=MGRAY, fontsize=10)
    axes[0].set_xticks([])

    axes[1].hist(clean, bins=28, color=INDIGO, edgecolor=WHITE, alpha=0.82)
    title_suffix = f"  ·  {n_out} outlier{'s' if n_out != 1 else ''} removed" if n_out > 0 else ""
    axes[1].set_title(f'Histogram (participants ≤ {PARTICIPANT_MAX:,}){title_suffix}', fontsize=11)
    axes[1].set_xlabel('Participants', color=MGRAY, fontsize=10)
    axes[1].set_ylabel('Events', color=MGRAY, fontsize=10)

    if n_out > 0:
        axes[1].axvline(
            PARTICIPANT_MAX, color=ROSE, linewidth=1.2, linestyle="--", alpha=0.7,
            label=f"Outlier threshold: {PARTICIPANT_MAX:,}",
        )
        axes[1].legend(fontsize=8, framealpha=0.9, edgecolor="#EBEBEB")

    plt.tight_layout()
    return fig


def c_heatmap(df):
    order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    heat = (
        df.groupby(['weekday', 'hour'])
        .size()
        .unstack(fill_value=0)
        .reindex(order, fill_value=0)
    )
    fig, ax = plt.subplots(figsize=(14, 4.5), facecolor=WHITE)
    ax.set_facecolor(WHITE)
    cmap = LinearSegmentedColormap.from_list(
        'indigo', [WHITE, '#EEF2FF', '#A5B4FC', INDIGO, '#312E81']
    )
    sns.heatmap(heat, ax=ax, cmap=cmap, linewidths=0.4, linecolor='#F8FAFC',
                cbar_kws={'label': 'Events', 'shrink': 0.8})
    ax.set_title('Event Density — Weekday × Start Hour', pad=14)
    ax.set_xlabel('Hour of Day', color=MGRAY, fontsize=10)
    ax.set_ylabel('', color=MGRAY)
    plt.tight_layout()
    return fig


def c_yearly_trend(df):
    yearly = df.groupby('year').agg(
        events=('event_name', 'count'),
        participants=('participants', 'sum'),
        hours=('duration_hours', 'sum'),
    ).reset_index()
    if len(yearly) < 2:
        return None
    fig, axes = plt.subplots(1, 3, figsize=(14, 4), facecolor=WHITE)
    yr_colors = CHART_COLORS[:len(yearly)]
    for ax in axes:
        ax.set_facecolor(WHITE)
        ax.spines[['top', 'right', 'left']].set_visible(False)
        ax.spines['bottom'].set_color('#EBEBEB')
        ax.tick_params(colors=MGRAY, length=0)
        ax.yaxis.grid(True, color='#F4F4F8', linewidth=0.7)
        ax.set_axisbelow(True)

    axes[0].bar(yearly['year'].astype(str), yearly['events'],
                color=yr_colors, edgecolor=WHITE, width=0.5, alpha=0.92)
    axes[0].set_title('Events per Year')
    for b in axes[0].patches:
        axes[0].text(b.get_x() + b.get_width() / 2, b.get_height() + 1,
                     f"{int(b.get_height()):,}", ha='center', fontsize=9, color=MGRAY, fontweight='500')

    axes[1].bar(yearly['year'].astype(str), yearly['participants'].fillna(0),
                color=yr_colors, edgecolor=WHITE, width=0.5, alpha=0.92)
    axes[1].set_title('Total Participants per Year')
    for b in axes[1].patches:
        axes[1].text(b.get_x() + b.get_width() / 2, b.get_height() + 1,
                     f"{int(b.get_height()):,}", ha='center', fontsize=9, color=MGRAY, fontweight='500')

    axes[2].bar(yearly['year'].astype(str), yearly['hours'].fillna(0),
                color=yr_colors, edgecolor=WHITE, width=0.5, alpha=0.92)
    axes[2].set_title('Total Hours per Year')
    for b in axes[2].patches:
        axes[2].text(b.get_x() + b.get_width() / 2, b.get_height() + 1,
                     f"{b.get_height():.0f}h", ha='center', fontsize=9, color=MGRAY, fontweight='500')

    plt.tight_layout()
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# MACHINE LEARNING — CANCELLATION PREDICTOR
# ══════════════════════════════════════════════════════════════════════════════

ML_FEATURES = ['hour', 'month', 'weekday_num', 'duration_hours',
               'room_enc', 'activity_enc', 'participants_clean']

WEEKDAY_MAP = {
    'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3,
    'Friday': 4, 'Saturday': 5, 'Sunday': 6,
}


def _prepare_ml_df(df: pd.DataFrame):
    d = df.copy()
    d = d[d['status'].isin(['Tenu', 'Annulé'])].copy()
    d['target'] = (d['status'] == 'Annulé').astype(int)
    d['weekday_num']        = d['weekday'].map(WEEKDAY_MAP).fillna(0).astype(int)
    d['participants_clean'] = d['participants'].clip(0, PARTICIPANT_MAX).fillna(0)
    d['duration_hours']     = d['duration_hours'].fillna(0).clip(0, 72)
    d['hour']               = d['hour'].fillna(0).astype(int)
    d['month']              = d['month'].fillna(1).astype(int)
    room_enc = LabelEncoder()
    act_enc  = LabelEncoder()
    d['room_enc']     = room_enc.fit_transform(d['room'].fillna('Unknown'))
    d['activity_enc'] = act_enc.fit_transform(d['activity_type'].fillna('Unknown'))
    X = d[ML_FEATURES].values
    y = d['target'].values
    return X, y, room_enc, act_enc, d


@st.cache_data(show_spinner=False)
def train_cancellation_model(file_bytes, file_name):
    df, *_ = load_and_clean(file_bytes, file_name)
    X, y, room_enc, act_enc, d = _prepare_ml_df(df)

    if len(np.unique(y)) < 2 or len(y) < 50:
        return None

    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

    candidates = {
        'Random Forest':       RandomForestClassifier(n_estimators=200, max_depth=8,
                                                      class_weight='balanced', random_state=42),
        'Gradient Boosting':   GradientBoostingClassifier(n_estimators=150, max_depth=4,
                                                          learning_rate=0.08, random_state=42),
        'Logistic Regression': Pipeline([('scaler', StandardScaler()),
                                         ('lr', LogisticRegression(class_weight='balanced',
                                                                    max_iter=500, random_state=42))]),
    }

    scores = {}
    for name, model in candidates.items():
        cv_scores = cross_val_score(model, X, y, cv=cv, scoring='f1', n_jobs=-1)
        scores[name] = cv_scores

    best_name  = max(scores, key=lambda k: scores[k].mean())
    best_model = candidates[best_name]
    best_model.fit(X, y)

    if hasattr(best_model, 'feature_importances_'):
        importances = best_model.feature_importances_
    elif hasattr(best_model, 'named_steps'):
        importances = best_model.named_steps['lr'].coef_[0]
    else:
        importances = np.zeros(len(ML_FEATURES))

    y_pred = best_model.predict(X)
    cm = confusion_matrix(y, y_pred)

    room_risk = (
        d.groupby('room')['target']
        .agg(total='count', cancelled='sum')
        .assign(cancel_rate=lambda x: (x['cancelled'] / x['total'] * 100).round(1))
        .sort_values('cancel_rate', ascending=False)
        .reset_index()
    )

    d['pred_proba'] = best_model.predict_proba(X)[:, 1]
    high_risk = (
        d[d['pred_proba'] >= 0.6]
        .groupby(['room', 'activity_type', 'weekday'])
        .size()
        .reset_index(name='count')
        .sort_values('count', ascending=False)
        .head(10)
    )

    return {
        'model':       best_model,
        'best_name':   best_name,
        'scores':      scores,
        'importances': importances,
        'cm':          cm,
        'room_risk':   room_risk,
        'high_risk':   high_risk,
        'room_enc':    room_enc,
        'act_enc':     act_enc,
        'X':           X,
        'y':           y,
        'n_samples':   len(y),
        'cancel_rate': round(y.mean() * 100, 1),
    }


def predict_single(model_result, hour, month, weekday, duration, room, activity, participants):
    room_enc = model_result['room_enc']
    act_enc  = model_result['act_enc']

    def safe_encode(enc, val):
        classes = list(enc.classes_)
        return classes.index(val) if val in classes else 0

    row = np.array([[
        hour,
        month,
        WEEKDAY_MAP.get(weekday, 0),
        min(duration, 72),
        safe_encode(room_enc, room),
        safe_encode(act_enc, activity),
        min(participants, PARTICIPANT_MAX),
    ]])
    proba = model_result['model'].predict_proba(row)[0][1]
    return round(proba * 100, 1)


def c_ml_feature_importance(importances, feature_names):
    fig, ax = make_fig(8, 4)
    idx = np.argsort(np.abs(importances))
    colors_bar = [ROSE if v > 0 else INDIGO for v in importances[idx]]
    ax.barh([feature_names[i] for i in idx], importances[idx],
            color=colors_bar, edgecolor=WHITE, height=0.55, alpha=0.9)
    ax.set_title('What influences cancellations the most?')
    ax.set_xlabel('Influence level (higher = more impact on the prediction)', color=MGRAY, fontsize=10)
    ax.axvline(0, color=MGRAY, linewidth=0.8, linestyle='--')
    plt.tight_layout()
    return fig


def c_ml_cv_scores(scores):
    fig, ax = make_fig(8, 4)
    names      = list(scores.keys())
    means      = [scores[n].mean() for n in names]
    stds       = [scores[n].std()  for n in names]
    colors_bar = [INDIGO, TEAL, AMBER]
    bars = ax.bar(names, means, yerr=stds, color=colors_bar[:len(names)],
                  edgecolor=WHITE, width=0.45, alpha=0.9,
                  capsize=6, error_kw=dict(color=MGRAY, linewidth=1.5))
    for b, m in zip(bars, means):
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.015,
                f'{round(m * 100)}%', ha='center', fontsize=10, fontweight='600', color=BLACK)
    ax.set_ylim(0, 1)
    ax.set_title('Which prediction method works best on your data?')
    ax.set_ylabel('Accuracy', color=MGRAY, fontsize=10)
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1))
    plt.tight_layout()
    return fig


def c_ml_confusion(cm):
    fig, ax = plt.subplots(figsize=(5, 4), facecolor=WHITE)
    ax.set_facecolor(WHITE)
    labels = [['TN', 'FP'], ['FN', 'TP']]
    cmap   = LinearSegmentedColormap.from_list('cm_cmap', [WHITE, '#C7D2FE', INDIGO])
    ax.imshow(cm, cmap=cmap, aspect='auto')
    for i in range(2):
        for j in range(2):
            ax.text(j, i, f'{labels[i][j]}\n{cm[i, j]:,}',
                    ha='center', va='center', fontsize=12,
                    fontweight='700', color=BLACK if cm[i, j] < cm.max() * 0.6 else WHITE)
    ax.set_xticks([0, 1]); ax.set_yticks([0, 1])
    ax.set_xticklabels(['Predicted: Held', 'Predicted: Cancelled'], fontsize=9, color=MGRAY)
    ax.set_yticklabels(['Actually: Held', 'Actually: Cancelled'], fontsize=9, color=MGRAY)
    ax.set_title('How often was it right?', pad=12)
    ax.spines[['top', 'right', 'left', 'bottom']].set_visible(False)
    plt.tight_layout()
    return fig


def c_ml_room_risk(room_risk):
    top = room_risk[room_risk['total'] >= 5].head(12)
    if top.empty:
        return None
    fig, ax = make_fig(9, max(3.5, len(top) * 0.6 + 1))
    max_r      = top['cancel_rate'].max()
    colors_bar = [ROSE if v == max_r else '#FDA4AF' for v in top['cancel_rate'].values]
    ax.barh(top['room'], top['cancel_rate'], color=colors_bar,
            edgecolor=WHITE, height=0.58, alpha=0.92)
    for p in ax.patches:
        ax.text(p.get_width() + 0.5, p.get_y() + p.get_height() / 2,
                f'{p.get_width():.1f}%', va='center', fontsize=9, color=MGRAY, fontweight='500')
    ax.xaxis.set_major_formatter(mticker.PercentFormatter())
    ax.set_title('Which rooms have the most cancellations? (historical)')
    ax.set_xlabel('% of events that were cancelled', color=MGRAY, fontsize=10)
    ax.invert_yaxis()
    plt.tight_layout()
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# CSS / DESIGN SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    background: #FAFAFA !important;
    color: #1A1A2E;
}

[data-testid="stSidebar"] {
    background: #1E1B4B !important;
    border-right: 1px solid #312E81 !important;
}
[data-testid="stSidebar"] > div { padding-top: 0 !important; }
[data-testid="stSidebar"] * { color: #E0E7FF !important; }
[data-testid="stSidebar"] hr { border-color: #312E81 !important; }
[data-testid="stSidebar"] .stFileUploader {
    background: #2D2A5E !important;
    border: 1.5px dashed #6366F1 !important;
    border-radius: 10px;
    padding: 14px;
}
[data-testid="stSidebar"] .stFileUploader label { color: #A5B4FC !important; }
[data-testid="stSidebar"] .stSelectbox > div > div {
    background: #2D2A5E !important;
    border-color: #4338CA !important;
    color: #E0E7FF !important;
}
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] .stCaption { color: #6366F1 !important; }

.stApp { background: #FAFAFA !important; }
.block-container { padding-top: 0 !important; padding-bottom: 2rem; max-width: 100% !important; }

.kpi-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 18px 16px 14px;
    border: 1px solid #F0F0F0;
    box-shadow: 0 1px 6px rgba(0,0,0,0.04);
    text-align: center;
}
.kpi-value {
    font-size: clamp(1.2rem, 2vw, 1.9rem);
    font-weight: 700;
    color: #1A1A2E;
    letter-spacing: -0.02em;
    line-height: 1;
}
.kpi-label {
    font-size: 0.63rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #94A3B8;
    margin-top: 6px;
}

[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 1.5px solid #EBEBEB !important;
    gap: 0;
}
[data-testid="stTabs"] [role="tab"] {
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
    color: #94A3B8 !important;
    border-radius: 0 !important;
    padding: 10px 22px !important;
    border: none !important;
    background: transparent !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    color: #4F46E5 !important;
    border-bottom: 2.5px solid #4F46E5 !important;
}

.stDownloadButton > button, .stButton > button {
    background: #4F46E5 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
    padding: 10px 22px !important;
    box-shadow: 0 2px 8px rgba(79,70,229,0.25) !important;
    transition: all 0.15s ease !important;
}
.stDownloadButton > button:hover, .stButton > button:hover {
    background: #4338CA !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 14px rgba(79,70,229,0.3) !important;
}

[data-testid="stSelectbox"] > div > div,
[data-testid="stMultiSelect"] > div > div {
    border-color: #C7D2FE !important;
    border-radius: 8px !important;
    background: #FFFFFF !important;
    color: #1A1A2E !important;
}

[data-testid="stMultiSelect"] span[data-baseweb="tag"] {
    background: #EEF2FF !important;
    border: 1px solid #C7D2FE !important;
    border-radius: 6px !important;
    color: #3730A3 !important;
    font-weight: 600 !important;
    font-size: 0.78rem !important;
}
[data-testid="stMultiSelect"] span[data-baseweb="tag"] span { color: #3730A3 !important; }
[data-testid="stMultiSelect"] span[data-baseweb="tag"] button svg { fill: #6366F1 !important; }
[data-testid="stMultiSelect"] span[data-baseweb="tag"] button:hover svg { fill: #3730A3 !important; }
[data-testid="stMultiSelect"] [data-baseweb="select"] > div { background: #FFFFFF !important; }

[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border-radius: 10px !important;
    padding: 16px 14px 12px !important;
    border: 1px solid #E2E8F0 !important;
    border-top: 3px solid #4F46E5 !important;
    box-shadow: 0 1px 6px rgba(0,0,0,0.04) !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Inter', sans-serif !important;
    font-size: 1.5rem !important;
    font-weight: 700 !important;
    color: #1A1A2E !important;
}
[data-testid="stMetricLabel"] {
    font-size: 0.67rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    color: #64748B !important;
}

[data-testid="stExpander"] {
    border: 1px solid #E2E8F0 !important;
    border-radius: 10px !important;
    background: #FFFFFF !important;
    margin-bottom: 8px !important;
}
[data-testid="stExpander"] summary {
    font-weight: 600 !important;
    font-size: 0.86rem !important;
    color: #374151 !important;
    padding: 12px 16px !important;
}
[data-testid="stExpander"] summary:hover {
    background: #F8F9FF !important;
    border-radius: 10px !important;
}

.section-label {
    font-size: 0.72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #4F46E5;
    margin: 2rem 0 1rem;
    display: flex;
    align-items: center;
    gap: 10px;
}
.section-label::after {
    content: '';
    flex: 1;
    height: 1px;
    background: #EBEBEB;
}

.chart-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 24px 24px 18px;
    margin-bottom: 16px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.04);
    border: 1px solid #F0F0F0;
}

.stAlert {
    border-radius: 10px !important;
    border-left: 3px solid #4F46E5 !important;
    background: #F0F0FF !important;
}

[data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden !important; }
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #F8F8F8; }
::-webkit-scrollbar-thumb { background: #E2E8F0; border-radius: 4px; }

.chat-bubble-user {
    background: #4F46E5;
    color: #FFFFFF;
    border-radius: 14px 14px 4px 14px;
    padding: 12px 16px;
    margin: 6px 0 6px auto;
    max-width: 72%;
    font-size: 0.88rem;
    line-height: 1.5;
    width: fit-content;
}
.chat-bubble-assistant {
    background: #FFFFFF;
    color: #1A1A2E;
    border-radius: 14px 14px 14px 4px;
    padding: 12px 16px;
    margin: 6px auto 6px 0;
    max-width: 80%;
    font-size: 0.88rem;
    line-height: 1.5;
    border: 1px solid #E2E8F0;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    width: fit-content;
}
.chat-bubble-assistant .thought {
    font-size: 0.74rem;
    color: #94A3B8;
    font-style: italic;
    margin-bottom: 6px;
    border-bottom: 1px solid #F0F0F0;
    padding-bottom: 6px;
}
.chat-error {
    background: #FFF1F2;
    border: 1px solid #FECDD3;
    border-radius: 10px;
    padding: 12px 16px;
    color: #E11D48;
    font-size: 0.82rem;
    margin: 6px 0;
}
.chat-code {
    background: #F8F9FF;
    border: 1px solid #E0E7FF;
    border-radius: 8px;
    padding: 10px 14px;
    font-size: 0.76rem;
    font-family: 'Courier New', monospace;
    color: #3730A3;
    margin-top: 6px;
    overflow-x: auto;
    white-space: pre-wrap;
}
.chat-input-area {
    position: sticky;
    bottom: 0;
    background: #FAFAFA;
    padding: 12px 0 4px;
    border-top: 1px solid #EBEBEB;
    margin-top: 16px;
}
.chat-suggestions {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 14px;
}
.chat-suggestion-pill {
    background: #EEF2FF;
    border: 1px solid #C7D2FE;
    border-radius: 20px;
    padding: 5px 12px;
    font-size: 0.76rem;
    font-weight: 600;
    color: #3730A3;
    cursor: pointer;
}
</style>
"""


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="Event Analytics",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown(CSS, unsafe_allow_html=True)
    chart_style()

    # ── SIDEBAR ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#4F46E5,#3730A3);
                    padding:28px 24px 24px;margin:-1rem -1rem 20px;border-bottom:1px solid #312E81;">
        <div style="font-family:'Inter',sans-serif;font-size:1.35rem;font-weight:700;
                    color:#FFFFFF;letter-spacing:-0.02em;line-height:1.2;">Event Analytics</div>
        <div style="font-size:0.7rem;font-weight:500;color:rgba(255,255,255,0.6);
                    letter-spacing:0.08em;text-transform:uppercase;margin-top:5px;">Pipeline · v4.2</div>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Upload data file", type=['csv', 'xlsx', 'xls'],
            help="data.csv or data.xlsx — columns are auto-detected",
        )

        st.markdown("<hr style='border-color:#EBEBEB;margin:20px 0'>", unsafe_allow_html=True)
        st.markdown("""
        <div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:0.12em;color:#818CF8;margin-bottom:10px;">
        Supported columns
        </div>
        <div style="font-size:0.78rem;color:#A5B4FC;line-height:2.1;">
        title · startTime · endTime<br>
        status · visibility<br>
        room / space · organization<br>
        participantNb · theme
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#EBEBEB;margin:20px 0'>", unsafe_allow_html=True)
        st.caption("Powered by Streamlit · matplotlib · openpyxl · Gemini")

    # ── HEADER BANNER ─────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:#FFFFFF;padding:24px 32px 22px;margin:-4rem -4rem 0;
                margin-bottom:28px;border-bottom:1px solid #EBEBEB;
                display:flex;align-items:center;gap:16px;">
    <div style="width:44px;height:44px;background:#4F46E5;
                border-radius:10px;display:flex;align-items:center;justify-content:center;
                flex-shrink:0;">
        <span style="font-size:1.2rem;color:white;">📊</span>
    </div>
    <div>
        <div style="font-family:'Inter',sans-serif;font-size:1.3rem;
                    font-weight:700;color:#1A1A2E;letter-spacing:-0.02em;line-height:1.2;">
        Event Analytics Pipeline
        </div>
        <div style="font-size:0.78rem;color:#94A3B8;margin-top:3px;">
        Upload · clean · analyse · export
        </div>
    </div>
    <div style="margin-left:auto;">
        <span style="background:#EEF2FF;color:#4F46E5;font-size:0.68rem;font-weight:600;
                    letter-spacing:0.06em;padding:4px 10px;border-radius:6px;
                    border:1px solid #C7D2FE;">v4.2</span>
    </div>
    </div>
    """, unsafe_allow_html=True)

    # ── NO FILE STATE ─────────────────────────────────────────────────────────
    if uploaded is None:
        st.markdown("""
        <div style="background:#FFFFFF;border-radius:14px;padding:52px;text-align:center;
                    border:1.5px dashed #E2E8F0;margin-top:20px;">
        <div style="font-size:3rem;margin-bottom:12px;">📂</div>
        <div style="font-family:'Inter',sans-serif;font-size:1.3rem;font-weight:700;
                    color:#1A1A2E;letter-spacing:-0.01em;">No data loaded</div>
        <div style="color:#94A3B8;margin-top:8px;font-size:0.9rem;">
            Upload a <strong>data.csv</strong> or <strong>data.xlsx</strong> in the sidebar to begin
        </div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4)
        for col, icon, label, desc, color in [
            (col1, "🧹", "Clean",   "Removes empty rows, dupes & fixes bad timestamps", EMERALD),
            (col2, "📊", "Analyse", "12+ charts auto-generated from your data",         INDIGO),
            (col3, "🔍", "Filter",  "Slice by status, type & year live",                AMBER),
            (col4, "💾", "Export",  "Styled Excel + removed rows report + CSV",         VIOLET),
        ]:
            col.markdown(f"""
            <div style="background:#FFFFFF;border-radius:12px;padding:24px;text-align:center;
                        border:1px solid #F0F0F0;box-shadow:0 1px 6px rgba(0,0,0,0.04);
                        margin-top:16px;">
            <div style="font-size:1.6rem;margin-bottom:10px;">{icon}</div>
            <div style="font-weight:700;font-size:0.85rem;color:{color};
                        letter-spacing:0.04em;margin-bottom:6px;">{label}</div>
            <div style="font-size:0.77rem;color:#94A3B8;line-height:1.5;">{desc}</div>
            </div>
            """, unsafe_allow_html=True)
        st.stop()

    # ── LOAD DATA ─────────────────────────────────────────────────────────────
    with st.spinner("Processing data…"):
        df, empty_rows, duplicate_rows, negative_rows = load_and_clean(
            uploaded.read(), uploaded.name
        )

    clean_participants = remove_outliers(df['participants'])
    n_outliers         = int(df['participant_outlier'].sum())

    total_events = len(df)
    held         = int((df['status'] == 'Tenu').sum())
    cancelled    = int((df['status'] == 'Annulé').sum())
    cancel_rate  = round(cancelled / total_events * 100, 1) if total_events else 0
    total_hours  = round(df['duration_hours'].sum(), 1)
    total_part   = int(clean_participants.sum(skipna=True))
    avg_part     = round(clean_participants.mean(skipna=True), 1)
    unique_orgs  = int(df['organization'].nunique())
    unique_rooms = int(df['room'].replace('', pd.NA).dropna().nunique())

    # ── KPI STRIP ─────────────────────────────────────────────────────────────
    st.markdown('<div class="section-label">Key Metrics</div>', unsafe_allow_html=True)

    kpis = [
        ("Total Events",  f"{total_events:,}",   "#4F46E5"),
        ("Held",          f"{held:,}",            "#059669"),
        ("Cancelled",     f"{cancelled:,}",       "#E11D48"),
        ("Cancel Rate",   f"{cancel_rate}%",      "#7C3AED"),
        ("Total Hours",   f"{total_hours:,.1f}h", "#0284C7"),
        ("Participants*", f"{total_part:,}",      "#0D9488"),
        ("Avg / Event*",  f"{avg_part:,.1f}",     "#D97706"),
        ("Organisations", f"{unique_orgs:,}",     "#DB2777"),
        ("Rooms",         f"{unique_rooms:,}",    "#1A1A2E"),
    ]

    cols = st.columns(len(kpis))
    for col, (label, value, color) in zip(cols, kpis):
        col.markdown(f"""
        <div class="kpi-card" style="border-top:3px solid {color};">
        <div class="kpi-value">{value}</div>
        <div class="kpi-label">{label}</div>
        </div>
        """, unsafe_allow_html=True)

    if n_outliers > 0:
        st.markdown(
            f'<div style="font-size:0.73rem;color:#94A3B8;margin-top:8px;">'
            f'* Participant metrics exclude <strong style="color:#E11D48">{n_outliers}</strong> '
            f'event{"s" if n_outliers != 1 else ""} with participants > {PARTICIPANT_MAX:,} (treated as outliers)</div>',
            unsafe_allow_html=True,
        )

    # ── FILTERS ───────────────────────────────────────────────────────────────
    st.markdown('<div class="section-label">Filters</div>', unsafe_allow_html=True)
    f1, f2, f3 = st.columns(3)
    with f1:
        sel_status = st.selectbox('Status', ['All'] + sorted(df['status'].dropna().unique().tolist()))
    with f2:
        sel_type = st.selectbox('Activity Type', ['All'] + sorted(df['activity_type'].dropna().unique().tolist()))
    with f3:
        years     = sorted(df['year'].dropna().unique().tolist())
        sel_years = st.multiselect('Year(s)', years, default=years)

    dff = df.copy()
    if sel_status != 'All': dff = dff[dff['status'] == sel_status]
    if sel_type   != 'All': dff = dff[dff['activity_type'] == sel_type]
    if sel_years:           dff = dff[dff['year'].isin(sel_years)]

    if len(dff) == 0:
        st.warning("No events match the selected filters.")
        st.stop()

    st.markdown(
        f'<div style="font-size:0.76rem;color:#94A3B8;margin-bottom:4px;">'
        f'Showing <strong style="color:#1A1A2E">{len(dff):,}</strong> of '
        f'<strong style="color:#1A1A2E">{total_events:,}</strong> events</div>',
        unsafe_allow_html=True,
    )

    # ── TABS ──────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "📊  Charts",
        "🗂  Data",
        "🧹  Cleaning",
        "📋  Statistics",
        "💾  Export",
        "🔮  Cancellation Predictor",
        "💬  Ask your data",
    ])

    # ── TAB 1: CHARTS ─────────────────────────────────────────────────────────
    with tab1:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_status(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col_b:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_weekday(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.pyplot(c_monthly(dff), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        yoy_fig = c_yearly_trend(dff)
        if yoy_fig:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(yoy_fig, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        col_c, col_d = st.columns(2)
        with col_c:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_activity(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col_d:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_hour(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        col_e, col_f = st.columns(2)
        with col_e:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_top_rooms(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col_f:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_top_orgs(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        col_g, col_h = st.columns(2)
        with col_g:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_avg_participants(dff), use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with col_h:
            cr_fig = c_cancel_rate(dff)
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            if cr_fig:
                st.pyplot(cr_fig, use_container_width=True)
            else:
                st.info("No cancellation data with current filters.")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.pyplot(c_room_hours(dff), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.pyplot(c_participants_dist(dff), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.pyplot(c_heatmap(dff), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── TAB 2: DATA ───────────────────────────────────────────────────────────
    with tab2:
        show_cols = [c for c in [
            'event_name', 'activity_type', 'start_time', 'end_time',
            'room', 'status', 'organization', 'participants',
            'duration_hours', 'participant_outlier',
        ] if c in dff.columns]
        st.dataframe(dff[show_cols], use_container_width=True, height=520)
        st.caption(
            f"{len(dff):,} rows · {len(show_cols)} columns shown · "
            "participant_outlier = True means excluded from participant stats"
        )

    # ── TAB 3: CLEANING ───────────────────────────────────────────────────────
    with tab3:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

        n_orig    = total_events + len(empty_rows) + len(duplicate_rows)
        n_empty   = len(empty_rows)
        n_dupes   = len(duplicate_rows)
        n_neg     = len(negative_rows)
        n_out_all = int(df['participant_outlier'].sum())

        cleaning_stats = [
            ("Original Rows",        f"{n_orig:,}",    "#4F46E5"),
            ("Empty Removed",        f"{n_empty:,}",   "#E11D48"),
            ("Duplicates Removed",   f"{n_dupes:,}",   "#D97706"),
            ("Neg Duration Fixed",   f"{n_neg:,}",     "#0D9488"),
            ("Participant Outliers", f"{n_out_all:,}", "#7C3AED"),
        ]
        c1, c2, c3, c4, c5 = st.columns(5)
        for col, (label, value, color) in zip([c1, c2, c3, c4, c5], cleaning_stats):
            col.markdown(f"""
            <div style="background:#FFFFFF;border-radius:10px;padding:18px 12px 14px;
                        border:1px solid #E2E8F0;border-top:3px solid {color};
                        box-shadow:0 1px 6px rgba(0,0,0,0.04);text-align:center;margin-bottom:16px;">
            <div style="font-family:'Inter',sans-serif;font-size:1.6rem;font-weight:700;
                        color:#1A1A2E;letter-spacing:-0.02em;line-height:1;">{value}</div>
            <div style="font-size:0.63rem;font-weight:600;text-transform:uppercase;
                        letter-spacing:0.09em;color:#94A3B8;margin-top:6px;">{label}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)

        if not empty_rows.empty:
            with st.expander(f"Empty rows  ·  {len(empty_rows):,}"):
                st.dataframe(empty_rows, use_container_width=True)
        if not duplicate_rows.empty:
            with st.expander(f"Duplicate rows  ·  {len(duplicate_rows):,}"):
                st.dataframe(duplicate_rows, use_container_width=True)
        if not negative_rows.empty:
            with st.expander(f"Negative-duration fixed  ·  {len(negative_rows):,}"):
                st.dataframe(negative_rows, use_container_width=True)
        outlier_rows = df[df['participant_outlier']]
        if not outlier_rows.empty:
            with st.expander(
                f"Participant outliers (> {PARTICIPANT_MAX:,}, kept but excluded from stats)  ·  {len(outlier_rows):,}"
            ):
                st.dataframe(
                    outlier_rows[['event_name', 'start_time', 'room', 'participants', 'organization']],
                    use_container_width=True,
                )
        multi = dff[dff['duration_flag'] == 'multi-day']
        if not multi.empty:
            with st.expander(f"Multi-day events kept  ·  {len(multi):,}"):
                st.dataframe(
                    multi[['event_name', 'start_time', 'end_time', 'duration_hours']],
                    use_container_width=True,
                )

    # ── TAB 4: STATISTICS ─────────────────────────────────────────────────────
    with tab4:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

        clean_dff    = remove_outliers(dff['participants'])
        total_part_f = clean_dff.sum(skipna=True)
        total_dur_f  = dff['duration_hours'].sum()
        avg_dur_f    = dff['duration_hours'].mean()
        avg_part_f   = clean_dff.mean(skipna=True)

        st.markdown(
            '<div class="section-label">Aggregate Totals — Filtered View</div>',
            unsafe_allow_html=True,
        )
        s1, s2, s3, s4 = st.columns(4)
        for col, val, label, color, suffix in [
            (s1, f"{int(total_part_f):,}", "Total Participants*",       TEAL,   ""),
            (s2, f"{total_dur_f:,.1f}",    "Total Hours Booked",        SKY,    "h"),
            (s3, f"{avg_part_f:.1f}",      "Avg Participants / Event*", AMBER,  ""),
            (s4, f"{avg_dur_f:.2f}",       "Avg Duration / Event",      VIOLET, "h"),
        ]:
            col.markdown(f"""
            <div style="background:#FFFFFF;border-radius:12px;padding:22px 16px 18px;
                        border-left:4px solid {color};border:1px solid #F0F0F0;
                        box-shadow:0 1px 6px rgba(0,0,0,0.04);text-align:center;margin-bottom:16px;">
            <div style="font-family:'Inter',sans-serif;font-size:2rem;font-weight:700;
                        color:#1A1A2E;letter-spacing:-0.02em;line-height:1;">
                {val}<span style="font-size:1rem;color:{color};margin-left:2px;">{suffix}</span>
            </div>
            <div style="font-size:0.63rem;font-weight:600;text-transform:uppercase;
                        letter-spacing:0.1em;color:#94A3B8;margin-top:8px;">{label}</div>
            </div>
            """, unsafe_allow_html=True)

        if n_outliers > 0:
            st.caption(f"* {n_outliers} event(s) with participants > {PARTICIPANT_MAX:,} are excluded from participant stats")

        st.markdown('<div class="section-label">Descriptive Statistics</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Duration (hours)**")
            st.dataframe(
                dff['duration_hours'].describe().rename('duration_hours').to_frame(),
                use_container_width=True,
            )
        with col2:
            st.markdown("**Participants (outliers excluded)**")
            st.dataframe(
                clean_dff.describe().rename('participants_clean').to_frame(),
                use_container_width=True,
            )

        st.markdown(
            '<div class="section-label">Events by Month × Activity Type</div>',
            unsafe_allow_html=True,
        )
        pivot = dff.groupby(['month_name', 'activity_type']).size().unstack(fill_value=0)
        st.dataframe(pivot, use_container_width=True)

        st.markdown(
            '<div class="section-label">Participants by Activity Type (outliers excluded)</div>',
            unsafe_allow_html=True,
        )
        clean_dff_full = dff[~dff['participant_outlier']]
        part_by_type   = clean_dff_full.groupby('activity_type')['participants'].agg(
            Total='sum', Average='mean', Max='max', Min='min', Count='count'
        ).round(1)
        part_by_type['Total'] = part_by_type['Total'].apply(lambda x: f"{int(x):,}")
        st.dataframe(part_by_type, use_container_width=True)

        st.markdown(
            '<div class="section-label">Duration by Room</div>',
            unsafe_allow_html=True,
        )
        dur_by_room = (
            dff[dff['room'] != '']
            .groupby('room')['duration_hours']
            .agg(Total_Hours='sum', Avg_Hours='mean', Events='count')
            .round(1)
            .sort_values('Total_Hours', ascending=False)
        )
        st.dataframe(dur_by_room, use_container_width=True)

    # ── TAB 5: EXPORT ─────────────────────────────────────────────────────────
    with tab5:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)

        for col, icon, title, desc, color, btn_label, btn_key, data_fn, filename, mime in [
            (
                col1, "📗", "Cleaned Events",
                "Styled Excel with dropdowns, table format & SUMMARY sheet",
                EMERALD, "Download cleaned_events.xlsx", "dl_clean",
                lambda: build_excel(dff), "cleaned_events.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                col2, "🗑️", "Removed Rows",
                "Empty rows · duplicates · negative durations",
                AMBER, "Download removed_rows.xlsx", "dl_removed",
                lambda: build_removed_excel(empty_rows, duplicate_rows, negative_rows),
                "removed_rows.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                col3, "📄", "Filtered CSV",
                "Current filtered view as plain CSV — ready for any tool",
                SKY, "Download filtered_events.csv", "dl_csv",
                lambda: dff.to_csv(index=False).encode(),
                "filtered_events.csv", "text/csv",
            ),
        ]:
            with col:
                st.markdown(f"""
                <div style="background:#FFFFFF;border-radius:12px;padding:24px 20px 20px;
                            text-align:center;box-shadow:0 1px 6px rgba(0,0,0,0.04);
                            border:1px solid #F0F0F0;border-top:3px solid {color};margin-bottom:8px;">
                <div style="font-size:1.8rem;margin-bottom:10px;">{icon}</div>
                <div style="font-weight:700;font-size:0.95rem;color:#1A1A2E;margin-bottom:8px;">{title}</div>
                <div style="font-size:0.78rem;color:#94A3B8;margin-bottom:18px;line-height:1.5;">{desc}</div>
                </div>
                """, unsafe_allow_html=True)
                with st.spinner("Building…"):
                    file_data = data_fn()
                st.download_button(
                    btn_label, data=file_data,
                    file_name=filename, mime=mime, key=btn_key,
                    use_container_width=True,
                )

    # ── TAB 6: ML PREDICTOR ───────────────────────────────────────────────────
    with tab6:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
        st.markdown('<div class="section-label">Will my event get cancelled?</div>', unsafe_allow_html=True)

        st.markdown("""
        <div style="background:#F0F0FF;border-radius:10px;padding:14px 18px;
                    border-left:3px solid #4F46E5;margin-bottom:20px;font-size:0.84rem;color:#374151;line-height:1.6;">
        This tool <strong>learned from your past event history</strong> to estimate whether a new event
        is likely to be cancelled. It looks for patterns — like which rooms or days tend to have more
        cancellations — and uses those to give you an early warning.
        The more events in your data, the more reliable the estimate.
        </div>
        """, unsafe_allow_html=True)

        with st.spinner("Analysing your past events…"):
            uploaded.seek(0)
            ml = train_cancellation_model(uploaded.read(), uploaded.name)

        if ml is None:
            st.warning("Not enough data to make predictions yet. You need at least 50 events with a Held or Cancelled status.")
        else:
            best_f1   = ml['scores'][ml['best_name']].mean()

            mk1, mk2, mk3, mk4 = st.columns(4)
            for col, val, label, color in [
                (mk1, "✅ Ready",                              "Prediction status",           INDIGO),
                (mk2, f"{round(best_f1 * 100)}% accurate",    "How often it's correct",      EMERALD),
                (mk3, f"{ml['n_samples']:,} events",          "Past events it learned from", TEAL),
                (mk4, f"{ml['cancel_rate']}% cancelled",      "Your overall cancellation rate", ROSE),
            ]:
                col.markdown(f"""
                <div style="background:#FFFFFF;border-radius:10px;padding:18px 12px 14px;
                            border:1px solid #E2E8F0;border-top:3px solid {color};
                            box-shadow:0 1px 6px rgba(0,0,0,0.04);text-align:center;margin-bottom:16px;">
                <div style="font-family:'Inter',sans-serif;font-size:1.1rem;font-weight:700;
                            color:#1A1A2E;line-height:1.2;">{val}</div>
                <div style="font-size:0.63rem;font-weight:600;text-transform:uppercase;
                            letter-spacing:0.09em;color:#94A3B8;margin-top:6px;">{label}</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("""
            <div style="font-size:0.82rem;color:#64748B;margin:8px 0 20px;line-height:1.6;">
            The charts below show <strong>which factors drive cancellations</strong> in your data,
            and how well the prediction compares to what actually happened.
            </div>
            """, unsafe_allow_html=True)

            ca1, ca2 = st.columns(2)
            with ca1:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.pyplot(c_ml_cv_scores(ml['scores']), use_container_width=True)
                st.markdown(
                    '<div style="font-size:0.75rem;color:#94A3B8;margin-top:8px;">'
                    'Three different methods were tested — the one with the highest bar was chosen.</div>',
                    unsafe_allow_html=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)
            with ca2:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.markdown(
                    '<div style="font-size:0.78rem;color:#64748B;margin-bottom:8px;">'
                    'The diagonal boxes show correct predictions. Bigger numbers on the diagonal = better.</div>',
                    unsafe_allow_html=True,
                )
                st.pyplot(c_ml_confusion(ml['cm']), use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

            feature_labels = [
                'Time of day', 'Month of year', 'Day of week', 'Event duration',
                'Room used', 'Type of activity', 'Number of guests',
            ]
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.pyplot(c_ml_feature_importance(ml['importances'], feature_labels),
                      use_container_width=True)
            st.markdown(
                '<div style="font-size:0.75rem;color:#94A3B8;margin-top:8px;">'
                'Longer bars = stronger influence on whether an event gets cancelled.</div>',
                unsafe_allow_html=True,
            )
            st.markdown('</div>', unsafe_allow_html=True)

            rr_fig = c_ml_room_risk(ml['room_risk'])
            if rr_fig:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.pyplot(rr_fig, use_container_width=True)
                st.markdown(
                    '<div style="font-size:0.75rem;color:#94A3B8;margin-top:8px;">'
                    'Only rooms with 5+ past events are shown.</div>',
                    unsafe_allow_html=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)

            if not ml['high_risk'].empty:
                st.markdown(
                    '<div class="section-label">⚠️ Combinations most likely to result in cancellation</div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<div style="font-size:0.82rem;color:#64748B;margin-bottom:12px;">'
                    'These room + activity + day combinations have the highest predicted cancellation risk '
                    'based on your historical data.</div>',
                    unsafe_allow_html=True,
                )
                st.dataframe(ml['high_risk'], use_container_width=True)

            st.markdown('<div class="section-label">Check a specific event</div>', unsafe_allow_html=True)

            st.markdown(
                '<div style="font-size:0.82rem;color:#64748B;margin-bottom:16px;line-height:1.6;">'
                "Fill in the details of an event you're planning. "
                'The tool will estimate — based on patterns from your past events — '
                'how likely it is to be cancelled. No technical knowledge needed.</div>',
                unsafe_allow_html=True,
            )

            known_rooms = sorted([r for r in ml['room_enc'].classes_ if r and r != 'Unknown'])
            known_activities = sorted([a for a in ml['act_enc'].classes_ if a and a != 'Unknown'])

            if "prediction_result" not in st.session_state:
                st.session_state.prediction_result = None

            with st.form("prediction_form"):
                p1, p2, p3 = st.columns(3)
                p4, p5, p6, p7 = st.columns(4)

                with p1:
                    pred_room = st.selectbox('Room', known_rooms)
                with p2:
                    pred_activity = st.selectbox('Activity type', known_activities)
                with p3:
                    pred_weekday = st.selectbox('Day of week',
                        ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])
                with p4:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#64748B;margin-bottom:2px;">'
                        '🕐 <strong>What time does it start?</strong><br>'
                        '<span style="font-size:0.72rem;">0 = midnight · 12 = noon · 17 = 5 PM</span></div>',
                        unsafe_allow_html=True,
                    )
                    pred_hour = st.slider('Start hour', 0, 23, 9, format='%d:00')
                with p5:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#64748B;margin-bottom:2px;">'
                        '📅 <strong>Which month?</strong><br>'
                        '<span style="font-size:0.72rem;">1 = January · 12 = December</span></div>',
                        unsafe_allow_html=True,
                    )
                    pred_month = st.slider('Month', 1, 12, 6)
                with p6:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#64748B;margin-bottom:2px;">'
                        '⏱️ <strong>How long is the event?</strong><br>'
                        '<span style="font-size:0.72rem;">In hours — e.g. 2.0 = two hours</span></div>',
                        unsafe_allow_html=True,
                    )
                    pred_duration = st.slider('Duration (hours)', 0.5, 24.0, 2.0, step=0.5, format='%.1f h')
                with p7:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#64748B;margin-bottom:2px;">'
                        '👥 <strong>How many people expected?</strong><br>'
                        f'<span style="font-size:0.72rem;">Your estimated number of attendees</span></div>',
                        unsafe_allow_html=True,
                    )
                    pred_participants = st.slider('Expected guests', 0, PARTICIPANT_MAX, 50)

                submitted = st.form_submit_button("🔮 Predict cancellation risk", use_container_width=True)

            if submitted:
                risk = predict_single(ml, pred_hour, pred_month, pred_weekday,
                                      pred_duration, pred_room, pred_activity, pred_participants)
                st.session_state.prediction_result = risk

            if st.session_state.prediction_result is not None:
                risk = st.session_state.prediction_result

                if risk >= 60:
                    risk_color = ROSE; risk_label = "⚠️ This event is at high risk of being cancelled"
                    risk_bg = "#FFF1F2"
                    risk_tip = "Consider choosing a different room or day, or following up with the organiser early."
                elif risk >= 35:
                    risk_color = AMBER; risk_label = "🟡 There's a moderate chance this gets cancelled"
                    risk_bg = "#FFFBEB"
                    risk_tip = "Keep an eye on this one — it may be worth a reminder closer to the date."
                else:
                    risk_color = EMERALD; risk_label = "✅ This event is unlikely to be cancelled"
                    risk_bg = "#F0FDF4"
                    risk_tip = "Looking good based on past patterns."

                st.markdown(
                    f"""<div style="background:{risk_bg};border-radius:14px;padding:28px;text-align:center;
                                border:2px solid {risk_color};margin-top:12px;">
                        <div style="font-size:0.85rem;font-weight:700;color:{risk_color};margin-bottom:8px;">
                            {risk_label}
                        </div>
                        <div style="font-size:3.5rem;font-weight:800;color:{risk_color};
                                    letter-spacing:-0.03em;line-height:1;">
                            {risk}%
                        </div>
                        <div style="font-size:0.82rem;color:#64748B;margin-top:10px;">
                            estimated cancellation probability · based on {ml['n_samples']:,} past events
                        </div>
                        <div style="font-size:0.8rem;color:{risk_color};margin-top:8px;font-style:italic;">
                            {risk_tip}
                        </div>
                    </div>""",
                    unsafe_allow_html=True,
                )

    # ── TAB 7: ASK YOUR DATA (Chat) ───────────────────────────────────────────
    with tab7:
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

       

        EXAMPLE_QUESTIONS = [
            "How many events were cancelled last year?",
            "Which room had the most events?",
            "What's the average number of participants per event?",
            "Which organisation booked the most events?",
            "Show me events with more than 200 participants",
            "What day of the week has the highest cancellation rate?",
            "How many unique organisations used the space?",
            "What's the longest event we've ever held?",
        ]

        if "chat_history_display" not in st.session_state:
            st.session_state.chat_history_display = []
        if "chat_history_ollama" not in st.session_state:
            st.session_state.chat_history_ollama = []
        if "chat_prefill" not in st.session_state:
            st.session_state.chat_prefill = ""

        st.markdown(
            '<div style="font-size:0.72rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:0.1em;color:#4F46E5;margin-bottom:8px;">Try asking…</div>',
            unsafe_allow_html=True,
        )
        pill_cols = st.columns(4)
        for i, q in enumerate(EXAMPLE_QUESTIONS):
            if pill_cols[i % 4].button(q, key=f"pill_{i}", use_container_width=True):
                st.session_state.chat_prefill = q
                st.rerun()

        st.markdown("<hr style='border-color:#EBEBEB;margin:16px 0'>", unsafe_allow_html=True)

        for msg in st.session_state.chat_history_display:
            if msg["role"] == "user":
                st.markdown(
                    f'<div class="chat-bubble-user">{msg["content"]}</div>',
                    unsafe_allow_html=True,
                )
            else:
                meta = msg.get("meta", {})
                if meta.get("error"):
                    st.markdown(
                        f'<div class="chat-error">⚠️ {meta["error"]}</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    thought_html = (
                        f'<div class="thought">💭 {meta["thought"]}</div>'
                        if meta.get("thought") else ""
                    )
                    st.markdown(
                        f'<div class="chat-bubble-assistant">'
                        f'{thought_html}'
                        f'<div>{msg["content"]}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    if isinstance(meta.get("result_df"), pd.DataFrame):
                        st.dataframe(meta["result_df"], use_container_width=True)
                    if meta.get("code"):
                        with st.expander("🔍 See generated code", expanded=False):
                            st.code(meta["code"], language="python")

        st.markdown('<div class="chat-input-area">', unsafe_allow_html=True)

        left_col, right_col = st.columns([5, 1])
        with left_col:
            user_input = st.text_input(
                "Ask a question about your data",
                value=st.session_state.chat_prefill,
                placeholder="e.g. Which room has the highest cancellation rate?",
                key="chat_input",
                label_visibility="collapsed",
            )
        with right_col:
            send_clicked = st.button("Ask →", key="chat_send", use_container_width=True)

        if st.session_state.chat_prefill and user_input == st.session_state.chat_prefill:
            st.session_state.chat_prefill = ""

        if st.session_state.chat_history_display:
            if st.button("🗑️ Clear conversation", key="chat_clear"):
                st.session_state.chat_history_display = []
                st.session_state.chat_history_ollama = []
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

        if (send_clicked or user_input) and user_input.strip():
            question = user_input.strip()

            last_user = next(
                (m for m in reversed(st.session_state.chat_history_display)
                if m["role"] == "user"), None,
            )
            if last_user and last_user["content"] == question:
                st.stop()

            st.session_state.chat_history_display.append(
                {"role": "user", "content": question}
            )

            with st.spinner("Thinking…"):
                result = text_to_pandas(
                    question=question,
                    df=dff,
                    history=st.session_state.chat_history_ollama,
                )

            if result.get("error"):
                st.session_state.chat_history_display.append({
                    "role": "assistant",
                    "content": "",
                    "meta": {
                        "error":     result["error"],
                        "thought":   result.get("thought", ""),
                        "code":      result.get("code", ""),
                        "result_df": None,
                    },
                })
            else:
                st.session_state.chat_history_display.append({
                    "role": "assistant",
                    "content": result["answer"],
                    "meta": {
                        "thought":   result.get("thought", ""),
                        "code":      result.get("code", ""),
                        "result_df": result.get("result_df"),
                        "error":     None,
                    },
                })
                st.session_state.chat_history_ollama.append(
                    {"role": "user", "content": question}
                )
                st.session_state.chat_history_ollama.append(
                    {"role": "assistant", "content": result["answer"]}
                )
                if len(st.session_state.chat_history_ollama) > 20:
                    st.session_state.chat_history_ollama = (
                        st.session_state.chat_history_ollama[-20:]
                    )

            st.rerun()


if __name__ == "__main__":
    main()